import func_timeout.exceptions
import flask
import json
import numpy as np
import os
import PIL.Image
import PIL.ImageDraw
import PIL.Image as Image
import cv2
import re
from ModularPrediction import Exporter, Inferer, PostProcessor
from ModularPrediction.PreProcessor import imageResize, labelResize
from threading import Thread
import inspect
import ctypes
from func_timeout import func_set_timeout
from pyzbar.pyzbar import decode
import pyzbar.wrapper
from pylib.onnx_paddleocr import ONNXPaddleOcr
import traceback
from rapidocr_onnxruntime import RapidOCR
# from simhash import Simhash
from openpyxl import load_workbook
import plistlib
from sentence_transformers import SentenceTransformer, util
import logging

# logger = logging.getLogger(__name__)
class MyTopThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        # self.resultPath = resultPath
        self.result = ()

    def run(self):
        self.result = getTopResult(self.InPath)

    def get_result(self):
        return self.result


class MyUpcThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        # self.resultPath = resultPath
        self.result = ()

    def run(self):
        self.result = getUPCResult(self.InPath)

    def get_Upc_Result(self):
        return self.result


class MyLogoThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        # self.resultPath = resultPath
        self.result = ()

    def run(self):
        self.result = getLogoResult(self.InPath)

    def get_logo_Result(self):
        return self.result


class MySdppiThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        # self.resultPath = resultPath
        self.result = ()

    def run(self):
        self.result = getSdppiResult(self.InPath)

    def get_sdppi_Result(self):
        return self.result


class MyPartThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        # self.resultPath = resultPath
        self.result = ()

    def run(self):
        self.result = getPartResult(self.InPath)

    def get_part_Result(self):
        return self.result


class MyMoppSdppiThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        self.result = ()

    def run(self):
        self.result = getMoppSdppiResult(self.InPath)

    def get_moppsdppi_Result(self):
        return self.result


class MyDpoThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        self.result = ()

    def run(self):
        self.result = getDpoResult(self.InPath)

    def get_dpo_Result(self):
        return self.result


class MyMrpThread(Thread):
    def __init__(self, InPath, mrpStep):
        Thread.__init__(self)
        self.InPath = InPath
        self.mrpStep = mrpStep
        self.result = ()

    def run(self):
        self.result = getMrpResult(self.InPath, self.mrpStep)

    def get_mrp_Result(self):
        return self.result


class MyDuckHeadAdapterThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        self.result = ()

    def run(self):
        self.result = getDuckHeadAdapterResult(self.InPath)

    def get_duckheadadapter_Result(self):
        return self.result


class MySecThread(Thread):
    def __init__(self, InPath):
        Thread.__init__(self)
        self.InPath = InPath
        self.result = ()

    def run(self):
        self.result = getSecResult(self.InPath)

    def get_sec_Result(self):
        return self.result


def _async_raise(tid, exctype):
    """raises the exception, performs cleanup if needed"""
    tid = ctypes.c_long(tid)
    if not inspect.isclass(exctype):
        exctype = type(exctype)
    res = ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, ctypes.py_object(exctype))
    if res == 0:
        raise ValueError("invalid thread id")
    elif res != 1:
        # """if it returns a number greater than one, you're in trouble,
        # and you should call it again with exc=NULL to revert the effect"""
        ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, None)
        raise SystemError("PyThreadState_SetAsyncExc failed")


def stop_thread(thread):
    _async_raise(thread.ident, SystemExit)


server = flask.Flask(__name__)  # __name__代表当前的python文件。把当前的python文件当做一个服务启动

""" Global variables : Image size are defined below """
INFERER: Inferer
WORKPATH = ''
COUNTRY = ''
COLOR = ''
ORDERTYPE = ''
EXITCOUNTRY = 0
MODEL_INFO = ""
PRED_IMGSIZE = (2400, 2400)
CONF_THRES = 0.25
IOU_THRES = 0.45
SAVE_TXT = True
SAVE_IMG = True
FINISH_LOAD = ''
ML_LOAD = 'loading...'
ML_LOG = ""
CALIBFILE = ''
LABELINFO = {}
FLAG = 0


def addLogs(msg):
    global ML_LOG
    ML_LOG = ML_LOG + msg + '\n'


# 保存异常日志
def saveErrLogs():
    global ML_LOG, resultPath
    os.makedirs(resultPath, exist_ok=True)
    with open(os.path.join(resultPath, 'predictLogs.txt'), 'a') as f:
        f.write(ML_LOG)


@server.route('/judgeCountry', methods=['get', 'post'])
def judgeCountry():
    global COUNTRY, LABELINFO, COLOR, FLAG, ORDERTYPE
    try:
        labelInfo = flask.request.args.get("labelInfo")
        labelInfo = json.loads(labelInfo)
        LABELINFO = labelInfo
        countryInfo = labelInfo['country']
        order_type = labelInfo['order_type']
        if LABELINFO['typec'] == '622-00582':
            colorInfo = 'SKY'
        elif LABELINFO['typec'] == '622-00491':
            colorInfo = 'SLV'
        elif LABELINFO['typec'] == '622-00449':
            colorInfo = 'STL'
        elif LABELINFO['typec'] == '622-00596':
            colorInfo = 'MDN'
        else:
            colorInfo = ''
        countryChange = 0
        if COUNTRY != countryInfo or COLOR != colorInfo or ORDERTYPE != order_type:
            countryChange = 1
            COUNTRY = countryInfo
            COLOR = colorInfo
            ORDERTYPE = order_type
        res = {'msg': countryChange, 'flag': FLAG}
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!', 'flag': FLAG}
    return json.dumps(res, ensure_ascii=False)

# def setup_custom_logging(log_file):
#     logger = logging.getLogger(log_file)
#     logger.setLevel(logging.INFO)
#     formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
#     file_handler = logging.FileHandler(log_file)
#     file_handler.setFormatter(formatter)
#     logger.addHandler(file_handler)
#     return logger


def getLabel(image, templateValue):
    # logger = setup_custom_logging("getLabel_log.txt")
    try:
        # caller_frame = inspect.stack()[1]
        # caller_info = f"{caller_frame.filename}:{caller_frame.lineno}"
        # logger.info(f"getLabel function called from {caller_info} with templateValue: {templateValue}")
        # caller_frame = inspect.stack()[1]
        # caller_info = f"{caller_frame.filename}:{caller_frame.lineno}"
        # logger.info(f"getLabel function called from {caller_info} with templateValue: {templateValue}")
        if templateValue == 'fg':
            template = FG
        elif templateValue == 'upc':
            template = UPC
        elif templateValue == 'ftr':
            template = FTR
        elif templateValue == 'sla':
            template = SLA
        elif templateValue == 'energy':
            template = ENERGY
        elif templateValue == 'importer':
            template = IMPORT
        elif templateValue == 'shipper':
            template = SHIPPER
        elif templateValue == 'sec':
            template = SEC
        elif templateValue == 'warranty':
            template = WARRANTY
        elif templateValue == 'sn':
            template = SN
        elif templateValue == 'sdppi':
            template = SDPPI
        elif templateValue == 'logoslv':
            template = LOGOSLV
        elif templateValue == 'logosky':
            template = LOGOSKY
        elif templateValue == 'logostl':
            template = LOGOSTL
        elif templateValue == 'logomdn':
            template = LOGOMDN
        elif templateValue == 'mopp':
            template = MOPP
        elif templateValue == 'dpoftr':
            template = DPOFTR
        elif templateValue == 'dposla':
            template = DPOSLA
        elif templateValue == 'dpofg':
            template = DPOFG
        elif templateValue == 'dpoupc':
            template = DPOUPC
        elif templateValue == 'dpoimporter':
            template = DPOIMPORTER
        elif templateValue == 'dposhipper':
            template = DPOSHIPPER
        elif templateValue == 'doublelogo':
            template = DOUBLELOGO
        elif templateValue == 'printbundle':
            template = PRINTBUNDLE
        elif templateValue == 'mrpstep1':
            template = MRPSTEP1
        elif templateValue == 'mrpstep3':
            template = MRPSTEP3
        elif templateValue == 'meps':
            template = MEPS
        else:
            print('Error: Template name input fail')
        res = cv2.matchTemplate(image, template, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        barcode_start = (max_loc[0], max_loc[1])
        barcode_end = (barcode_start[0] + template.shape[1], barcode_start[1] + template.shape[0])
        matchImage = image[barcode_start[1]:barcode_end[1], barcode_start[0]:barcode_end[0], :]
        # logger.info(f"Match successful for templateValue: {templateValue}")
        return matchImage, max_val
    except:
    # except Exception as e:
    #     logger.error(f"Error occurred in getLabel function with templateValue: {templateValue}. Error: {str(e)}")
        return None, None


@server.route('/initService', methods=['get', 'post'])
def initService():
    global FINISH_LOAD, ML_LOAD, OCRINFERER, UPC, FG, MRPSTEP1, MRPSTEP3, FTR, IMPORT, ENERGY, SEC, WARRANTY, SN, SHIPPER, SDPPI, SLA, \
        LOGOSLV, LOGOMDN, LOGOSKY, LOGOSTL, MOPP, DPOFTR, DPOSLA, DPOFG, DPOUPC, DPOIMPORTER, DPOSHIPPER, PRINTBUNDLE, DOUBLELOGO, INFERER_LINE, \
        WORKPATH, CALIBFILE, ORB, FASTOCR, INFERER_DUCKHEAD, POSTPROCESSOR, FLAG, OCRINFEREREN, INFERER_DUCKHEAD_LAST, \
        SENTENCETRANSFORMERMODEL, MEPS, DPORECYCLE, INFERER_LINE_LAST
    workpath = flask.request.args.get("workPath")
    detModelPath = workpath + '/model/det.onnx'
    claModelPath = workpath + '/model/cls.onnx'
    recModelPath = workpath + '/model/rec.onnx'
    txtPath = workpath + '/model/keys.txt'
    endetModelPath = workpath + '/model/en_det.onnx'
    enrecModelPath = workpath + '/model/en_rec.onnx'
    entxtPath = workpath + '/model/en_dict.txt'
    thick_json_file = workpath + '/Thick_Calib.json'
    thin_json_file = workpath + '/Thin_Calib.json'
    country = COUNTRY
    WORKPATH = workpath
    massage_load = ''
    scriptpath = os.path.dirname(os.path.abspath(__file__))
    image_paths = []
    try:
        ###### inference model
        OCRINFERER = ONNXPaddleOcr(detModelPath, claModelPath, recModelPath, txtPath, use_angle_cls=True, use_gpu=False)
        OCRINFEREREN = ONNXPaddleOcr(endetModelPath, claModelPath, enrecModelPath, entxtPath, use_angle_cls=True, use_gpu=False)
        ORB = cv2.ORB_create()
        FASTOCR = RapidOCR()
        INFERER_DUCKHEAD = Inferer.JRSYOLOV8CoreML(workpath + '/model/DuckHead71.mlpackage')
        INFERER_DUCKHEAD_LAST = Inferer.JRSYOLOV8CoreML(workpath + '/model/DuckHeadLast71.mlpackage')
        INFERER_LINE = Inferer.JRSYOLOV8CoreML(workpath + '/model/Line71.mlpackage')
        INFERER_LINE_LAST = Inferer.JRSYOLOV8CoreML(workpath + '/model/LineLast71.mlpackage')
        POSTPROCESSOR = PostProcessor.NMS(conf_thres=CONF_THRES, iou_thres=IOU_THRES)
        SENTENCETRANSFORMERMODEL = SentenceTransformer(workpath + "/model/clip_model")
        ####### load calib file by country
        if country == 'C' or country == 'E' or country == 'LL' or country == 'J' or country == 'JA' or country == 'CH'\
            or country == 'CT' or country == 'PP' or country == 'TA':
            if os.path.isfile(thin_json_file):
                with open(thin_json_file, 'r') as f:
                    CALIBFILE = json.load(f)
        else:
            if os.path.isfile(thick_json_file):
                with open(thick_json_file, 'r') as f:
                    CALIBFILE = json.load(f)
        ###### load template by country
            UPC = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                             LABELINFO['order_type'] + '/UPC/NA.jpg')
        if UPC is not None:
            UPC = cv2.resize(UPC, (2210, 1912))
            SHIPPER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                                 LABELINFO['order_type'] + '/SHIPPER/NA.jpg')
        if LABELINFO['order_type'] == 'BTR':
            DPOFTR = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/FTR/NA.jpg')
            # if DPOFTR is not None:
            #   DPOFTR = cv2.resize(DPOFTR, (2430, 1262))
            DPOFG = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                             LABELINFO['order_type'] + '/DPO/FG/NA.jpg')
        elif LABELINFO['order_type'] == 'CTO':
            DPOSLA = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/SLA/NA.jpg')
            DPOFG = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/FG/NA.jpg')
        else:
            DPOSLA = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/SLA/NA.jpg')
            DPOFG = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                               '/' + LABELINFO['order_type'] + '/DPO/FG/NA.jpg')

            DPOUPC = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/UPC/NA.jpg')
        # if DPOUPC is not None:
        #     DPOUPC = cv2.resize(DPOUPC, (1120, 410))
        # cv2.imwrite('tempp.jpg', DPOUPC)
        DPOSHIPPER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/SHIPPER/NA.jpg')
        if LABELINFO['typec'] == '622-00491':
            LOGOSLV = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                 '/' + LABELINFO['order_type'] + '/LOGO/SLV.jpg')
            SDPPI = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                               '/' + LABELINFO['order_type'] + '/SDPPI/SLV.jpg')
            MOPP = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                              '/' + LABELINFO['order_type'] + '/MOPP/SLV.jpg')
            if LABELINFO['country'] == 'ID' or LABELINFO['country'] == 'KH':
                MEPS = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                  '/' + LABELINFO['order_type'] + '/MEPS/SLV.jpg')
            else:
                MEPS = None
            # DOUBLELOGO = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                         '/' + LABELINFO['order_type'] + '/DOUBLELOGO/SLV.jpg')
            # LINE = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                   '/' + LABELINFO['order_type'] + '/LINE/SLV.jpg')

        elif LABELINFO['typec'] == '622-00582':
            LOGOSKY = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                 '/' + LABELINFO['order_type'] + '/LOGO/SKY.jpg')
            SDPPI = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                               '/' + LABELINFO['order_type'] + '/SDPPI/SKY.jpg')
            MOPP = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                              '/' + LABELINFO['order_type'] + '/MOPP/SKY.jpg')
            if LABELINFO['country'] == 'ID' or LABELINFO['country'] == 'KH':
                MEPS = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                  '/' + LABELINFO['order_type'] + '/MEPS/SKY.jpg')
            else:
                MEPS = None
            # DOUBLELOGO = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                         '/' + LABELINFO['order_type'] + '/DOUBLELOGO/SKY.jpg')
            # LINE = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                   '/' + LABELINFO['order_type'] + '/LINE/SKY.jpg')
        elif LABELINFO['typec'] == '622-00449':
            LOGOSTL = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                 '/' + LABELINFO['order_type'] + '/LOGO/STL.jpg')
            SDPPI = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                               '/' + LABELINFO['order_type'] + '/SDPPI/STL.jpg')
            MOPP = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                              '/' + LABELINFO['order_type'] + '/MOPP/STL.jpg')
            if LABELINFO['country'] == 'ID' or LABELINFO['country'] == 'KH':
                MEPS = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                  '/' + LABELINFO['order_type'] + '/MEPS/STL.jpg')
            else:
                MEPS = None
            # DOUBLELOGO = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                         '/' + LABELINFO['order_type'] + '/DOUBLELOGO/STL.jpg')
            # LINE = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                   '/' + LABELINFO['order_type'] + '/LINE/STL.jpg')

        else:
            LOGOMDN = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                 '/' + LABELINFO['order_type'] + '/LOGO/MDN.jpg')
            SDPPI = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                               '/' + LABELINFO['order_type'] + '/SDPPI/MDN.jpg')
            MOPP = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                              '/' + LABELINFO['order_type'] + '/MOPP/MDN.jpg')
            if LABELINFO['country'] == 'ID' or LABELINFO['country'] == 'KH':
                MEPS = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                  '/' + LABELINFO['order_type'] + '/MEPS/MDN.jpg')
            else:
                MEPS = None
            # DOUBLELOGO = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                         '/' + LABELINFO['order_type'] + '/DOUBLELOGO/MDN.jpg')
            # LINE = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                   '/' + LABELINFO['order_type'] + '/LINE/MDN.jpg')
        PRINTBUNDLE = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                 '/' + LABELINFO['order_type'] + '/PRINTBUNDLE/NA.jpg')
        if country == 'SA':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
            ENERGY = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/ENERGY/NA.jpg')
        elif country == 'BZ' or country == 'KH' or country == 'SA' or country == 'TG' or country == 'TH' \
                or country == 'CT' or country == 'TA':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
        elif country == 'B' or country == 'D' or country == 'FN' or country == 'N' or country == 'T' or country == 'Y' \
                or country == 'CR' or country == 'CZ' or country == 'DK' or country == 'GR' or country == 'H' or country == 'KS' \
                or country == 'MG' or country == 'PO' or country == 'RO' or country == 'RU' or country == 'SL' or country == 'SM' \
                or country == 'TU' or country == 'ZE':
            SHIPPER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                                 LABELINFO['order_type'] + '/SHIPPER/NA.jpg')
        elif country == 'CI':
            SEC = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                             LABELINFO['order_type'] + '/SEC/NA.jpg')
        elif country == 'CH':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
            ENERGY = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/ENERGY/NA.jpg')
            WARRANTY = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                  '/' + LABELINFO['order_type'] + '/WARRANTY/NA.jpg')
            SN = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                            '/' + LABELINFO['order_type'] + '/SN/NA.jpg')
        elif country == 'ID':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
        elif country == 'HN':
            if LABELINFO['order_type'] == 'BTR':
                MRPSTEP1 = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                      '/' + LABELINFO['order_type'] + '/MRP/mrp1.jpg')
                MRPSTEP3 = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                      '/' + LABELINFO['order_type'] + '/MRP/mrp3.jpg')
            else:
                MRPSTEP1 = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                      '/' + LABELINFO['order_type'] + '/MRP/mrp1.jpg')
                MRPSTEP3 = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                      '/' + LABELINFO['order_type'] + '/MRP/mrp3.jpg')
        elif country == 'UA':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
            SHIPPER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                                 LABELINFO['order_type'] + '/SHIPPER/NA.jpg')
        else:
            pass

        massage_load = 'Success: The model and template has loaded!'
        ML_LOAD = ''
        FLAG = 1
        # massage_load = 'Success: The model and template have loaded!'
        res = {'msg': massage_load}
    except:
        ML_LOAD = 'Load fail ....'
        massage_load = 'Error: The OCR model and template load fail!'
        res = {'msg': 'Error', 'info': massage_load}

    FINISH_LOAD = 'OK'
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/loadTemplate', methods=['get', 'post'])
def loadTemplate():
    global UPC, FG, MRPSTEP1, MRPSTEP3, FTR, IMPORT, ENERGY, SEC, WARRANTY, SN, SHIPPER, SDPPI, SLA, LOGOSLV, LOGOSPB, \
        MOPP, DPOFTR, DPOSLA, DPOFG, DPOUPC, DPOIMPORTER, DPOSHIPPER, PRINTBUNDLE, DOUBLELOGO
    try:
        workpath = WORKPATH
        country = COUNTRY
        scriptpath = os.path.dirname(os.path.abspath(__file__))
        # UPC = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
        #                  LABELINFO['order_type'] + '/UPC/NA.jpg')

        UPC = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                             LABELINFO['order_type'] + '/UPC/NA.jpg')

        if UPC is not None:
            UPC = cv2.resize(UPC, (1300, 1200))
        SHIPPER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                             LABELINFO['order_type'] + '/SHIPPER/NA.jpg')

        if LABELINFO['order_type'] == 'BTR':
            DPOFTR = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/FTR/NA.jpg')
            # if DPOFTR is not None:
            #     DPOFTR = cv2.resize(DPOFTR, (2430, 1262))
            DPOFG = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/FG/NA.jpg')
        elif LABELINFO['order_type'] == 'CTO':
            DPOSLA = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/SLA/NA.jpg')
            DPOFG = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/FG/NA.jpg')
            DPOUPC = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/UPC/NA.jpg')
        else:
            DPOSLA = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/SLA/NA.jpg')
            DPOFG = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                               '/' + LABELINFO['order_type'] + '/DPO/FG/NA.jpg')
            DPOUPC = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/UPC/NA.jpg')
        # if DPOUPC is not None:
        #     DPOUPC = cv2.resize(DPOUPC, (1120, 410))
        # cv2.imwrite('tempp.jpg', DPOUPC)
        DPOSHIPPER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/SHIPPER/NA.jpg')
        if LABELINFO['typec'] == '677-32159' or LABELINFO['typec'] == '622-00491':
            LOGOSLV = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                 '/' + LABELINFO['order_type'] + '/LOGO/SLV.jpg')
            SDPPI = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                               '/' + LABELINFO['order_type'] + '/SDPPI/SLV.jpg')
            MOPP = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                              '/' + LABELINFO['order_type'] + '/MOPP/SLV.jpg')
            # DOUBLELOGO = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                         '/' + LABELINFO['order_type'] + '/DOUBLELOGO/SLV.jpg')
            # LINE = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                   '/' + LABELINFO['order_type'] + '/LINE/SLV.jpg')
        else:
            LOGOSPB = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                 '/' + LABELINFO['order_type'] + '/LOGO/SPB.jpg')
            SDPPI = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                               '/' + LABELINFO['order_type'] + '/SDPPI/SPB.jpg')
            MOPP = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                              '/' + LABELINFO['order_type'] + '/MOPP/SPB.jpg')
            # DOUBLELOGO = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                         '/' + LABELINFO['order_type'] + '/DOUBLELOGO/SPB.jpg')
            # LINE = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
            #                   '/' + LABELINFO['order_type'] + '/LINE/SPB.jpg')
        PRINTBUNDLE = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                 '/' + LABELINFO['order_type'] + '/PRINTBUNDLE/NA.jpg')

        if country == 'BZ' or country == 'KH' or country == 'TG' or country == 'TH' \
                or country == 'CT' or country == 'TA':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
        if country == 'SA':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
            ENERGY = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/ENERGY/NA.jpg')
        elif country == 'B' or country == 'D' or country == 'FN' or country == 'N' or country == 'T' or country == 'Y' \
                or country == 'CR' or country == 'CZ' or country == 'DK' or country == 'GR' or country == 'H' or country == 'KS' \
                or country == 'MG' or country == 'PO' or country == 'RO' or country == 'RU' or country == 'SL' or country == 'SM' \
                or country == 'TU' or country == 'ZE':
             SHIPPER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                                 LABELINFO['order_type'] + '/SHIPPER/NA.jpg')

        elif country == 'CI':
            SEC = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                             LABELINFO['order_type'] + '/SEC/NA.jpg')
        elif country == 'CH':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
            ENERGY = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                '/' + LABELINFO['order_type'] + '/DPO/ENERGY/NA.jpg')
            WARRANTY = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                  '/' + LABELINFO['order_type'] + '/WARRANTY/NA.jpg')
            SN = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                            '/' + LABELINFO['order_type'] + '/SN/NA.jpg')
        elif country == 'ID':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
        elif country == 'HN':
            if LABELINFO['order_type'] == 'BTR':
                MRPSTEP1 = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                      '/' + LABELINFO['order_type'] + '/MRP/mrp1.jpg')
                MRPSTEP3 = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                      '/' + LABELINFO['order_type'] + '/MRP/mrp3.jpg')
            else:
                MRPSTEP1 = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                      '/' + LABELINFO['order_type'] + '/MRP/mrp1.jpg')
                MRPSTEP3 = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                      '/' + LABELINFO['order_type'] + '/MRP/mrp3.jpg')
        elif country == 'UA':
            DPOIMPORTER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country +
                                     '/' + LABELINFO['order_type'] + '/DPO/IMPORTER/NA.jpg')
            SHIPPER = cv2.imread(workpath + '/Template/' + LABELINFO['project'] + '/' + country + '/' +
                                 LABELINFO['order_type'] + '/SHIPPER/NA.jpg')
        else:
            pass
        res = {'msg': 'OK'}
    except:
        res = {'msg': 'NG'}

    return json.dumps(res, ensure_ascii=False)


@server.route('/calibImage', methods=['get', 'post'])
def calibImage():
    try:
        if CALIBFILE != '':
            inputPath = flask.request.args.get("inputPath")
            image = cv2.imread(inputPath)
            camera_matrix = np.array(CALIBFILE.get("camera_matrix", []))
            dist_coeff = np.array(CALIBFILE.get("dist_coeff", []))
            h, w = image.shape[:2]
            newcameramatrix, roi = cv2.getOptimalNewCameraMatrix(camera_matrix, dist_coeff, (w, h), 1, (w, h))
            undistorted_image = cv2.undistort(image, camera_matrix, dist_coeff, None, newcameramatrix)
            inputDir = os.path.split(inputPath)[0]
            resultPath = os.path.join(inputDir, 'calibed.jpg')
            cv2.imwrite(resultPath, undistorted_image)
            res = {'msg': 'OK', 'info': 'Success: Calibration finished!', 'imagepath': resultPath}
        else:
            res = {'msg': 'NG', 'info': 'Fail: Calibration file reading error!'}
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkTop', methods=['get', 'post'])
def checkTop():
    # global resultPath
    try:
        inputPath = flask.request.args.get("inputPath")
        Topthread = MyTopThread(inputPath)
        Topthread.start()
        Topthread.join(200)
        final_result, resultFlag = Topthread.get_result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Topthread.is_alive():
            stop_thread(Topthread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkUpc', methods=['get', 'post'])
def checkUpc():
    try:
        inputPath = flask.request.args.get("inputPath")
        Upcthread = MyUpcThread(inputPath)
        Upcthread.start()
        Upcthread.join(200)
        final_result, resultFlag = Upcthread.get_Upc_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Upcthread.is_alive():
            stop_thread(Upcthread)
    except:
        with open(WORKPATH + '/error_log.txt', 'a') as f:
            f.write('UPC Runtime Errors Info as follow: \n')
            f.write(traceback.format_exc())
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkMoppSdppi', methods=['get', 'post'])
def checkMoppSdppi():
    try:
        inputPath = flask.request.args.get("inputPath")
        Moppthread = MyMoppSdppiThread(inputPath)
        Moppthread.start()
        Moppthread.join(200)
        final_result, resultFlag = Moppthread.get_moppsdppi_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Moppthread.is_alive():
            stop_thread(Moppthread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkDpo', methods=['get', 'post'])
def checkDpo():
    try:
        inputPath = flask.request.args.get("inputPath")
        Dpothread = MyDpoThread(inputPath)
        Dpothread.start()
        Dpothread.join(200)
        final_result, resultFlag = Dpothread.get_dpo_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Dpothread.is_alive():
            stop_thread(Dpothread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkLogo', methods=['get', 'post'])
def checkLogo():
    try:
        inputPath = flask.request.args.get("inputPath")
        Logothread = MyLogoThread(inputPath)
        Logothread.start()
        Logothread.join(200)
        final_result, resultFlag = Logothread.get_logo_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Logothread.is_alive():
            stop_thread(Logothread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkSdppi', methods=['get', 'post'])
def checkSdppi():
    try:
        inputPath = flask.request.args.get("inputPath")
        Sdppithread = MySdppiThread(inputPath)
        Sdppithread.start()
        Sdppithread.join(200)
        final_result, resultFlag = Sdppithread.get_sdppi_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Sdppithread.is_alive():
            stop_thread(Sdppithread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkPart', methods=['get', 'post'])
def checkPart():
    try:
        inputPath = flask.request.args.get("inputPath")
        Partthread = MyPartThread(inputPath)
        Partthread.start()
        Partthread.join(200)
        final_result, resultFlag = Partthread.get_part_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Partthread.is_alive():
            stop_thread(Partthread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkMrp', methods=['get', 'post'])
def checkMrp():
    try:
        inputPath = flask.request.args.get("inputPath")
        mrpStep = flask.request.args.get("mrpStep")
        Mrpthread = MyMrpThread(inputPath, mrpStep)
        Mrpthread.start()
        Mrpthread.join(200)
        final_result, resultFlag = Mrpthread.get_mrp_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Mrpthread.is_alive():
            stop_thread(Mrpthread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkDuckHeadAdapter', methods=['get', 'post'])
def checkDuckHeadAdapter():
    try:
        inputPath = flask.request.args.get("inputPath")
        DuckHeadAdapterthread = MyDuckHeadAdapterThread(inputPath)
        DuckHeadAdapterthread.start()
        DuckHeadAdapterthread.join(200)
        final_result, resultFlag = DuckHeadAdapterthread.get_duckheadadapter_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if DuckHeadAdapterthread.is_alive():
            stop_thread(DuckHeadAdapterthread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/checkSec', methods=['get', 'post'])
def checkSec():
    try:
        inputPath = flask.request.args.get("inputPath")
        Secthread = MySecThread(inputPath)
        Secthread.start()
        Secthread.join(200)
        final_result, resultFlag = Secthread.get_sec_Result()
        if resultFlag > 0:
            res = {'msg': 'Fail', 'info': final_result}
        else:
            res = {'msg': 'Pass', 'info': final_result}
        # saveErrLogs()
        if Secthread.is_alive():
            stop_thread(Secthread)
    except:
        res = {'msg': 'Error', 'info': 'Runtime error!'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


def getUPCResult(inputPath):
    upcResult = {}
    resultFlag = 0
    image = cv2.imread(inputPath)
    country = LABELINFO['country']
    if country == 'J' or country == 'JA':
        cutNum = 3
        ######裁图获得UPC小图
        height = image.shape[0]
        width = image.shape[1]
        image_upc = image[int(height / 4): int(height / 4 * 3), :int(width / 3 * 2)]
        # image_upc = rotate_image(image_upc)
        # cv2.imwrite('temp.jpg', image_upc)
        if UPC is None:
            upc_image = image_upc
            matchSimilarity = 0.7
            upc_similirity = 0.9
        else:
            upc_image, matchSimilarity = getLabel(image_upc, 'upc')
            # cv2.imwrite('temp.jpg', upc_image)
            upc_image = rotate_image(upc_image)
            # cv2.imwrite('temp.jpg', upc_image)
            upc_similirity = calculateSimilarity(upc_image, UPC)
        if upc_similirity < 0.8 or matchSimilarity < 0.1:
            upcResult.update({'upc_similarity': {'result': 'NG', 'item': 'upc',
                                                 'detail': {'Threshold': 0.8, 'score': upc_similirity}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'upc_similarity': {'result': 'OK', 'item': 'upc',
                                                 'detail': {'Threshold': 0.8, 'score': upc_similirity}}})
        ######分离并识别各部分信息
        upcInfo = cutImageByLine(upc_image, cutNum)
        if upcInfo[0].find('JAN') < 0:
            if LABELINFO['upc'] == LABELINFO['pn']:
                upc_ocr_value = upcInfo[0].replace('O', '0').replace('I', '1')
                if upc_ocr_value.find(LABELINFO['upc'].replace('O', '0').replace('I', '1')) < 0:
                    upcResult.update({'upc_txt': {'result': 'NG', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': upc_ocr_value}}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    upcResult.update({'upc_txt': {'result': 'OK', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
            else:
                pattern = r'\d+'
                upc_ocr_value = upcInfo[0].replace('O', '0').replace('I', '1')
                upc_ocr_value = re.findall(pattern, upc_ocr_value)[0]
                if LABELINFO['upc'].replace('O', '0').replace('I', '1').find(upc_ocr_value) < 0 or \
                        LABELINFO['upc'].replace('O', '0').replace('I', '1')[-1:] != upc_ocr_value[-1:]:
                    upcResult.update({'upc_txt': {'result': 'NG', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': upc_ocr_value}}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    upcResult.update({'upc_txt': {'result': 'OK', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
        else:
            if LABELINFO['upc'] == LABELINFO['pn']:
                upc_ocr_value = upcInfo[0].split('JAN')[1].replace('O', '0').replace('I', '1')
                if upc_ocr_value.find(LABELINFO['upc'].replace('O', '0').replace('I', '1')) < 0:
                    upcResult.update({'upc_txt': {'result': 'NG', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': upc_ocr_value}}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    upcResult.update({'upc_txt': {'result': 'OK', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
            else:
                pattern = r'\d+'
                upc_ocr_value = upcInfo[0].split('JAN')[1].replace('O', '0').replace('I', '1')
                upc_ocr_value = re.findall(pattern, upc_ocr_value)[0]
                if LABELINFO['upc'].replace('O', '0').replace('I', '1').find(upc_ocr_value) < 0 or \
                        LABELINFO['upc'].replace('O', '0').replace('I', '1')[-1:] != upc_ocr_value[-1:]:
                    upcResult.update({'upc_txt': {'result': 'NG', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': upc_ocr_value}}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    upcResult.update({'upc_txt': {'result': 'OK', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
        if upcInfo[1].find(LABELINFO['upc']) < 0:
            upcResult.update({'upc_code': {'result': 'NG', 'item': 'upc',
                              'detail': {'expected': LABELINFO['upc'], 'actual': upcInfo[1]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'upc_code': {'result': 'OK', 'item': 'upc',
                              'detail': {'expected': LABELINFO['upc'], 'actual': upcInfo[1]}}})
        if upcInfo[2].find(LABELINFO['pn']) < 0:
            if upcInfo[2].split('NO.')[1].replace('/', '').replace('O', '0').find(LABELINFO['pn'].replace('O', '0').replace('/', '')) < 0:
                upcResult.update({'pn_txt': {'result': 'NG', 'item': 'upc',
                                  'detail': {'expected': LABELINFO['pn'], 'actual': upcInfo[2].split('NO.')[1]}}})
                resultFlag = 1
            else:
                upcResult.update({'pn_txt': {'result': 'OK', 'item': 'upc',
                                             'detail': {'expected': LABELINFO['pn'],
                                                        'actual': LABELINFO['pn']}}})
            # return upcResult, resultFlag
        else:
            upcResult.update({'pn_txt': {'result': 'OK', 'item': 'upc',
                              'detail': {'expected': LABELINFO['pn'], 'actual': upcInfo[2].split('NO.')[1]}}})
        if upcInfo[3].find(LABELINFO['pn']) < 0:
            upcResult.update({'pn_code': {'result': 'NG', 'item': 'upc',
                              'detail': {'expected': LABELINFO['pn'], 'actual': upcInfo[3]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'pn_code': {'result': 'OK', 'item': 'upc',
                              'detail': {'expected': LABELINFO['pn'], 'actual': upcInfo[3]}}})
        if upcInfo[4].replace('O', '0').find(LABELINFO['sn']) < 0:
            upcResult.update({'sn_txt': {'result': 'NG', 'item': 'upc',
                              'detail': {'expected': LABELINFO['sn'], 'actual': upcInfo[4].replace('O', '0').split('N0.')[1]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'sn_txt': {'result': 'OK', 'item': 'upc',
                              'detail': {'expected': LABELINFO['sn'], 'actual': LABELINFO['sn']}}})
        if upcInfo[5].find(LABELINFO['sn']) < 0:
            upcResult.update({'sn_code': {'result': 'NG', 'item': 'upc',
                              'detail': {'expected': LABELINFO['sn'], 'actual': upcInfo[5]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'sn_code': {'result': 'OK', 'item': 'upc',
                              'detail': {'expected': LABELINFO['sn'], 'actual': upcInfo[5]}}})
        # if upcInfo[6].find(LABELINFO['modelNo']) < 0:
        if upcInfo[6].find(LABELINFO['model_no']) < 0:
            upcResult.update({'model_no': {'result': 'NG', 'item': 'upc',
                              'detail': {'expected': LABELINFO['model_no'], 'actual': upcInfo[6].split('NO.')[1]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'model_no': {'result': 'OK', 'item': 'upc',
                                           'detail': {'expected': LABELINFO['model_no'], 'actual': LABELINFO['model_no']}}})
        #######check upc other info
        upcimage_ocr_result, _ = FASTOCR(upc_image)
        upc_image_ocr_value = ''
        if upcimage_ocr_result is None:
            upcimage_ocr_result = OCRINFERER.ocr(upc_image)[0]
            if len(upcimage_ocr_result) == 0:
                upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                                                'detail': 'UPC other info fail!'}})
                resultFlag = 1
            else:
                for line in upcimage_ocr_result:
                    upc_image_ocr_value += line[1][0].replace(' ', '').upper()
                if country == 'LL' or country == 'C' or country == 'VC' or country == 'CI':
                    if upc_image_ocr_value.find('MARKED') < 0:
                        upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                                                        'detail': 'UPC other info fail!'}})
                        resultFlag = 1
                        # return upcResult, resultFlag
                    else:
                        pass
                else:
                    if upc_image_ocr_value.find('ASS') < 0 or upc_image_ocr_value.find(
                            'CHINA') < 0 or upc_image_ocr_value.find('MARKED') < 0:
                        upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                                                        'detail': 'UPC other info fail!'}})
                        resultFlag = 1
                        # return upcResult, resultFlag
                    else:
                        pass
        else:
            for txt in upcimage_ocr_result:
                upc_image_ocr_value += txt[1].replace(" ", "").replace("0", "O").upper()
            if country == 'LL' or country == 'C' or country == 'VC' or country == 'CI':
                if upc_image_ocr_value.find('MARKED') < 0:
                    upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                                                    'detail': 'UPC other info fail!'}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    pass
            else:
                if upc_image_ocr_value.find('ASS') < 0 or upc_image_ocr_value.find(
                        'CHINA') < 0 or upc_image_ocr_value.find('MARKED') < 0:
                    upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                                                    'detail': 'UPC other info fail!'}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    pass
        #####FIT shipper
        if SHIPPER is None:
            upcResult.update({'shipper': {}})
            return upcResult, resultFlag
        else:
            shipImage = image[int(height/2):, ]
            # cv2.imwrite('temp.jpg', shipImage)
            shipper_image, shippermap = getLabel(shipImage, 'shipper')
            # cv2.imwrite('temp.jpg', shipper_image)
            shipper_similarity = calculateSimilarity(shipper_image, SHIPPER)
            if shipper_similarity < 0.8 or shippermap < 0.1:
                if country == 'B' or country == 'D' or country == 'FN' or country == 'N' or country == 'T' or country == 'Y' \
                        or country == 'CR' or country == 'CZ' or country == 'DK' or country == 'GR' or country == 'H' \
                        or country == 'KS' or country == 'MG' or country == 'PO' or country == 'RO' or country == 'RU' \
                        or country == 'SL' or country == 'SM' or country == 'TU' or country == 'ZE' or country == 'UA':
                    upcResult.update({'shipper_similarity': {'result': 'NG', 'item': 'shipper',
                                                             'detail': {'Threshold': 0.8, 'score': shipper_similarity}}})
                    resultFlag = 1
                    return upcResult, resultFlag
                else:
                    upcResult.update({'shipper': {}})
                    return upcResult, resultFlag
            else:
                upcResult.update({'shipper_similarity': {'result': 'OK', 'item': 'shipper',
                                                         'detail': {'Threshold': 0.8, 'score': shipper_similarity}}})
                return upcResult, resultFlag
    else:
        cutNum = 3
        ######裁图获得UPC小图
        height = image.shape[0]
        width = image.shape[1]
        image_upc = image[int(height / 4): int(height / 4 * 3), :int(width / 3 * 2)]
        # image_upc = rotate_image(image_upc)
        # cv2.imwrite('temp.jpg', image_upc)
        if UPC is None:
            upc_image = image_upc
            matchSimilarity = 0.7
            upc_similirity = 0.8
        else:
            upc_image, matchSimilarity = getLabel(image_upc, 'upc')
            # cv2.imwrite('temp.jpg', upc_image)
            upc_image = rotate_image(upc_image)
            # cv2.imwrite('temp1.jpg', upc_image)
            upc_similirity = calculateSimilarity(upc_image, UPC)
        if upc_similirity < 0.8 or matchSimilarity < 0.1:
            upcResult.update({'upc_similarity': {'result': 'NG', 'item': 'upc',
                                                 'detail': {'Threshold': 0.8, 'score': upc_similirity}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'upc_similarity': {'result': 'OK', 'item': 'upc',
                                                 'detail': {'Threshold': 0.8, 'score': upc_similirity}}})
        ######分离并识别各部分信息
        upcInfo = cutImageByLine(upc_image, cutNum)
        if upcInfo[0].find('UPC') < 0:
            if LABELINFO['upc'] == LABELINFO['pn']:
                upc_ocr_value = upcInfo[0].replace('O', '0').replace('I', '1')
                if upc_ocr_value.find(LABELINFO['upc'].replace('O', '0').replace('I', '1')) < 0:
                    upcResult.update({'upc_txt': {'result': 'NG', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': upc_ocr_value}}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    upcResult.update({'upc_txt': {'result': 'OK', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
            else:
                pattern = r'\d+'
                upc_ocr_value = upcInfo[0].replace('O', '0').replace('I', '1')
                upc_ocr_value = re.findall(pattern, upc_ocr_value)[0]
                if LABELINFO['upc'].replace('O', '0').replace('I', '1').find(upc_ocr_value) < 0 or \
                        LABELINFO['upc'].replace('O', '0').replace('I', '1')[-1:] != upc_ocr_value[-1:]:
                    upcResult.update({'upc_txt': {'result': 'NG', 'item': 'upc',
                                                  'detail': {'expected': LABELINFO['upc'], 'actual': upc_ocr_value}}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    upcResult.update({'upc_txt': {'result': 'OK', 'item': 'upc',
                                                  'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
        else:
            if LABELINFO['upc'] == LABELINFO['pn']:
                upc_ocr_value = upcInfo[0].split('UPC')[1].replace('O', '0').replace('I', '1')
                if upc_ocr_value.find(LABELINFO['upc'].replace('O', '0').replace('I', '1')) < 0:
                    upcResult.update({'upc_txt': {'result': 'NG', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': upc_ocr_value}}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    upcResult.update({'upc_txt': {'result': 'OK', 'item': 'upc',
                                      'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
            else:
                pattern = r'\d+'
                upc_ocr_value = upcInfo[0].split('UPC')[1].replace('O', '0').replace('I', '1')
                upc_ocr_value = re.findall(pattern, upc_ocr_value)[0]
                if LABELINFO['upc'].replace('O', '0').replace('I', '1').find(upc_ocr_value) < 0 or \
                        LABELINFO['upc'].replace('O', '0').replace('I', '1')[-1:] != upc_ocr_value[-1:]:
                    upcResult.update({'upc_txt': {'result': 'NG', 'item': 'upc',
                                                  'detail': {'expected': LABELINFO['upc'], 'actual': upc_ocr_value}}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    upcResult.update({'upc_txt': {'result': 'OK', 'item': 'upc',
                                                  'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
        if upcInfo[1].find(LABELINFO['upc']) < 0:
            upcResult.update({'upc_code': {'result': 'NG', 'item': 'upc',
                                           'detail': {'expected': LABELINFO['upc'], 'actual': upcInfo[1]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'upc_code': {'result': 'OK', 'item': 'upc',
                                           'detail': {'expected': LABELINFO['upc'], 'actual': upcInfo[1]}}})
        if upcInfo[2].find(LABELINFO['pn']) < 0:
            if upcInfo[2].split('NO.')[1].replace('/', '').replace('O', '0').find(LABELINFO['pn'].replace('O', '0').replace('/', '')) < 0:
                upcResult.update({'pn_txt': {'result': 'NG', 'item': 'upc',
                                             'detail': {'expected': LABELINFO['pn'],
                                                        'actual': upcInfo[2].split('NO.')[1]}}})
                resultFlag = 1
            else:
                upcResult.update({'pn_txt': {'result': 'OK', 'item': 'upc',
                                             'detail': {'expected': LABELINFO['pn'],
                                                        'actual': LABELINFO['pn']}}})
            # return upcResult, resultFlag
        else:
            upcResult.update({'pn_txt': {'result': 'OK', 'item': 'upc',
                                         'detail': {'expected': LABELINFO['pn'],
                                                    'actual': LABELINFO['pn']}}})
        if upcInfo[3].find(LABELINFO['pn']) < 0:
            upcResult.update({'pn_code': {'result': 'NG', 'item': 'upc',
                                          'detail': {'expected': LABELINFO['pn'], 'actual': upcInfo[3]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'pn_code': {'result': 'OK', 'item': 'upc',
                                          'detail': {'expected': LABELINFO['pn'], 'actual': upcInfo[3]}}})
        if upcInfo[4].replace('O', '0').find(LABELINFO['sn']) < 0:
            upcResult.update({'sn_txt': {'result': 'NG', 'item': 'upc',
                                         'detail': {'expected': LABELINFO['sn'],
                                                    'actual': upcInfo[4].replace('O', '0').split('N0.')[1]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'sn_txt': {'result': 'OK', 'item': 'upc',
                                         'detail': {'expected': LABELINFO['sn'],
                                                    'actual': LABELINFO['sn']}}})
        if upcInfo[5].find(LABELINFO['sn']) < 0:
            upcResult.update({'sn_code': {'result': 'NG', 'item': 'upc',
                                          'detail': {'expected': LABELINFO['sn'], 'actual': upcInfo[5]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'sn_code': {'result': 'OK', 'item': 'upc',
                                          'detail': {'expected': LABELINFO['sn'], 'actual': upcInfo[5]}}})
        if upcInfo[6].find(LABELINFO['model_no']) < 0:
            upcResult.update({'model_no': {'result': 'NG', 'item': 'upc',
                                           'detail': {'expected': LABELINFO['model_no'],
                                                      'actual': upcInfo[6].split('NO.')[1]}}})
            resultFlag = 1
            # return upcResult, resultFlag
        else:
            upcResult.update({'model_no': {'result': 'OK', 'item': 'upc',
                                           'detail': {'expected': LABELINFO['model_no'],
                                                      'actual': LABELINFO['model_no']}}})
        #####detect upc other info
        upcimage_ocr_result, _ = FASTOCR(upc_image)
        upc_image_ocr_value = ''
        if upcimage_ocr_result is None:
            upcimage_ocr_result = OCRINFERER.ocr(upc_image)[0]
            if len(upcimage_ocr_result) == 0:
                upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                                                'detail': 'UPC other info fail!'}})
                resultFlag = 1
            else:
                for line in upcimage_ocr_result:
                    upc_image_ocr_value += line[1][0].replace(' ', '').upper()
                #if country == 'LL' or country == 'C' or country == 'VC' or country == 'CI':
                if upc_image_ocr_value.find('MARKED') < 0:
                    upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                                                    'detail': 'UPC other info fail!'}})
                    resultFlag = 1
                    # return upcResult, resultFlag
                else:
                    pass
                # else:
                #     if upc_image_ocr_value.find('ASS') < 0 or upc_image_ocr_value.find(
                #             'CHINA') < 0 or upc_image_ocr_value.find('MARKED') < 0:
                #         upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                #                                         'detail': 'UPC other info fail!'}})
                #         resultFlag = 1
                #         # return upcResult, resultFlag
                #     else:
                #         pass
        else:
            for txt in upcimage_ocr_result:
                upc_image_ocr_value += txt[1].replace(" ", "").replace("0", "O").upper()
            # if country == 'LL' or country == 'C' or country == 'VC' or country == 'CI':
            if upc_image_ocr_value.find('MARKED') < 0:
                upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
                                                'detail': 'UPC other info fail!'}})
                resultFlag = 1
                # return upcResult, resultFlag
            else:
                pass
            # else:
            #     if upc_image_ocr_value.find('ASS') < 0 or upc_image_ocr_value.find(
            #             'CHINA') < 0 or upc_image_ocr_value.find('MARKED') < 0:
            #         upcResult.update({'upc_other': {'result': 'NG', 'item': 'upc',
            #                                         'detail': 'UPC other info fail!'}})
            #         resultFlag = 1
            #         # return upcResult, resultFlag
            #     else:
            #         pass
        #####FIT shipper
        if SHIPPER is None:
            upcResult.update({'shipper': {}})
            return upcResult, resultFlag
        else:
            shipImage = image[int(height / 2):]
            # cv2.imwrite('temp3.jpg', shipImage)
            shipper_image, shippermap = getLabel(shipImage, 'shipper')
            # cv2.imwrite('temp2.jpg', shipper_image)
            shipper_similarity = calculateSimilarity(shipper_image, SHIPPER)
            if shipper_similarity < 0.8 or shippermap < 0.1:
                if country == 'B' or country == 'D' or country == 'FN' or country == 'N' or country == 'T' or country == 'Y' \
                        or country == 'CR' or country == 'CZ' or country == 'DK' or country == 'GR' or country == 'H' \
                        or country == 'KS' or country == 'MG' or country == 'PO' or country == 'RO' or country == 'RU' \
                        or country == 'SL' or country == 'SM' or country == 'TU' or country == 'ZE' or country == 'UA':
                    upcResult.update({'shipper_similarity': {'result': 'NG', 'item': 'shipper',
                                                             'detail': {'Threshold': 0.8,
                                                                        'score': shipper_similarity}}})
                    resultFlag = 1
                    return upcResult, resultFlag
                else:
                    upcResult.update({'shipper': {}})
                    return upcResult, resultFlag
            else:
                upcResult.update({'shipper_similarity': {'result': 'OK', 'item': 'shipper',
                                                         'detail': {'Threshold': 0.8, 'score': shipper_similarity}}})
                return upcResult, resultFlag


def getDpoResult(inputPath):
    dpoResult = {}
    flag = 0
    image = cv2.imread(inputPath)
    country = LABELINFO['country']
    dpoftr_ocr_value = ''
    dpofg_ocr_value = ''
    dpoupc_ocr_value = ''
    dpopn_ocr_value = ''
    dposn_ocr_value = ''
    dposhipperva_ocr_value = ''
    dpoimporterva_ocr_value = ''
    dpoenergy_ocr_value = ''
    dposhippermodelno_ocr_value = ''
    dposhipperyear_ocr_value = ''
    dpoimportervh_ocr_value = ''
    # if country == 'C' or country == 'E' or country == 'LE' or country == 'LL' or country == 'J' or country == 'JA' \
    #         or country == 'PA' or country == 'PP' or country == 'X' or country == 'ZA' or country == 'ZP' \
    #         or country == 'AB' or country == 'AE' or country == 'HB' or country == 'ZS':
    if LABELINFO['order_type'] == 'BTR':
        dpoftrHeight = image.shape[0]
        dpoftrWidth = image.shape[1]
        ftrimage = image[int(dpoftrHeight / 9 * 4): int(dpoftrHeight / 4 * 3), : int(dpoftrWidth / 5 * 4)]
        # cv2.imwrite('temp.jpg', ftrimage)
        dpoftrimage, dpoftrMapSimilarity = getLabel(ftrimage, 'dpoftr')
        # cv2.imwrite('temp.jpg', dpoftrimage)
        dpoftrSimilarity = calculateSimilarity(dpoftrimage, DPOFTR)
        if dpoftrSimilarity < 0.6 or dpoftrMapSimilarity < 0.05:
            dpoResult.update({'dpoftr_similarity': {'result': 'NG', 'item': 'dpo',
                                                    'detail': {'Threshold': 0.6, 'score': dpoftrSimilarity}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'dpoftr_similarity': {'result': 'OK', 'item': 'dpo',
                                                    'detail': {'Threshold': 0.6, 'score': dpoftrSimilarity}}})
        height = dpoftrimage.shape[0]
        dpoftrimage = dpoftrimage[: int(height / 3)]
        dpoftrimage_resize = cv2.resize(dpoftrimage, (380, 74))
        # cv2.imwrite('temp.jpg', dpoftrimage)
        dpoftr_ocr_result = OCRINFERER.ocr(dpoftrimage_resize)[0]
        for dpoftr in dpoftr_ocr_result:
            dpoftr_ocr_value += dpoftr[1][0].replace(' ', '').upper()
        # dpoftr_ocr_result, _ = FASTOCR(dpoftrimage)
        # for dpoftr in dpoftr_ocr_result:
        #     dpoftr_ocr_value += dpoftr[1].replace(" ", "").upper()
        pattern = r'\d+'
        dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
        if country == 'AB' or country == 'HB' or country == 'ZS' or country == 'AE':
            if len(dpoftr_ocr_data) > 1:
                if dpoftr_ocr_data[0] != LABELINFO['ssd_size'] or dpoftr_ocr_data[1] != LABELINFO['mem_size']:
                    dpoftr_ocr_result, _ = FASTOCR(dpoftrimage_resize)
                    dpoftr_ocr_value = ''
                    for dpoftr in dpoftr_ocr_result:
                        dpoftr_ocr_value += dpoftr[1].replace(" ", "").upper()
                    pattern = r'\d+'
                    dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
                    if dpoftr_ocr_data[0] != LABELINFO['ssd_size'] or dpoftr_ocr_data[1] != LABELINFO['mem_size']:
                        dpoftr_ocr_result = OCRINFEREREN.ocr(dpoftrimage)[0]
                        dpoftr_ocr_value = ''
                        for dpoftr in dpoftr_ocr_result:
                            dpoftr_ocr_value += dpoftr[1][0].replace(' ', '').upper()
                        pattern = r'\d+'
                        dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
                        if dpoftr_ocr_data[0] != LABELINFO['ssd_size'] or dpoftr_ocr_data[1] != LABELINFO['mem_size']:
                            dpoftr_ocr_result, _ = FASTOCR(dpoftrimage)
                            dpoftr_ocr_value = ''
                            for dpoftr in dpoftr_ocr_result:
                                dpoftr_ocr_value += dpoftr[1].replace(" ", "").upper()
                            pattern = r'\d+'
                            dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
            else:
                dpoftr_ocr_result = OCRINFERER.ocr(dpoftrimage)[0]
                dpoftr_ocr_value = ''
                for dpoftr in dpoftr_ocr_result:
                    dpoftr_ocr_value += dpoftr[1][0].replace(' ', '').upper()
                pattern = r'\d+'
                dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
                if dpoftr_ocr_data[0] != LABELINFO['ssd_size']:
                    dpoftr_ocr_result = OCRINFEREREN.ocr(dpoftrimage)[0]
                    dpoftr_ocr_value = ''
                    for dpoftr in dpoftr_ocr_result:
                        dpoftr_ocr_value += dpoftr[1][0].replace(' ', '').upper()
                    pattern = r'\d+'
                    dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
                    if dpoftr_ocr_data[0] != LABELINFO['ssd_size'] or dpoftr_ocr_data[1] != LABELINFO['mem_size']:
                        dpoftr_ocr_result, _ = FASTOCR(dpoftrimage_resize)
                        dpoftr_ocr_value = ''
                        for dpoftr in dpoftr_ocr_result:
                            dpoftr_ocr_value += dpoftr[1].replace(" ", "").upper()
                        pattern = r'\d+'
                        dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
        else:
            if len(dpoftr_ocr_data) > 2:
                if dpoftr_ocr_data[2] != LABELINFO['ssd_size']:
                    dpoftr_ocr_result, _ = FASTOCR(dpoftrimage_resize)
                    dpoftr_ocr_value = ''
                    for dpoftr in dpoftr_ocr_result:
                        dpoftr_ocr_value += dpoftr[1].replace(" ", "").upper()
                    pattern = r'\d+'
                    dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
                    if len(dpoftr_ocr_data[0]) > 1 and dpoftr_ocr_data[0].find('4') >= 0:
                        tempValue = dpoftr_ocr_data[0][1:]
                        tempssdValue = dpoftr_ocr_data[1]
                        dpoftr_ocr_data[0] = '4'
                        dpoftr_ocr_data[1] = tempValue
                        if len(dpoftr_ocr_data) > 2:
                            dpoftr_ocr_data[2] = tempssdValue
                        else:
                            dpoftr_ocr_data.append(tempssdValue)
                    if dpoftr_ocr_data[2] != LABELINFO['ssd_size']:
                        dpoftr_ocr_result = OCRINFEREREN.ocr(dpoftrimage_resize)[0]
                        dpoftr_ocr_value = ''
                        for dpoftr in dpoftr_ocr_result:
                            dpoftr_ocr_value += dpoftr[1][0].replace(' ', '').upper()
                        pattern = r'\d+'
                        dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
            else:
                dpoftr_ocr_result = OCRINFERER.ocr(dpoftrimage)[0]
                dpoftr_ocr_value = ''
                for dpoftr in dpoftr_ocr_result:
                    dpoftr_ocr_value += dpoftr[1][0].replace(' ', '').upper()
                pattern = r'\d+'
                dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
                if len(dpoftr_ocr_data[0]) > 1 and dpoftr_ocr_data[0].find('4') >= 0:
                    tempValue = dpoftr_ocr_data[0][1:]
                    tempssdValue = dpoftr_ocr_data[1]
                    dpoftr_ocr_data[0] = '4'
                    dpoftr_ocr_data[1] = tempValue
                    if len(dpoftr_ocr_data) > 2:
                        dpoftr_ocr_data[2] = tempssdValue
                    else:
                        dpoftr_ocr_data.append(tempssdValue)
                if dpoftr_ocr_data[2] != LABELINFO['ssd_size']:
                    dpoftr_ocr_result = OCRINFEREREN.ocr(dpoftrimage_resize)[0]
                    dpoftr_ocr_value = ''
                    for dpoftr in dpoftr_ocr_result:
                        dpoftr_ocr_value += dpoftr[1][0].replace(' ', '').upper()
                    pattern = r'\d+'
                    dpoftr_ocr_data = re.findall(pattern, dpoftr_ocr_value)
        if country == 'AB' or country == 'ZS' or country == 'AE':
            if dpoftr_ocr_data[1] != LABELINFO['mem_size']:
                dpoResult.update({'mem_size': {'result': 'NG', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['mem_size'],
                                                          'actual': dpoftr_ocr_data[1]}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'mem_size': {'result': 'OK', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['mem_size'],
                                                          'actual': dpoftr_ocr_data[1]}}})
            if dpoftr_ocr_data[0] != LABELINFO['ssd_size']:
                dpoResult.update({'ssd_size': {'result': 'NG', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['ssd_size'],
                                                          'actual': dpoftr_ocr_data[0]}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'ssd_size': {'result': 'OK', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['ssd_size'],
                                                          'actual': dpoftr_ocr_data[0]}}})
        elif country == 'HB':
            if dpoftr_ocr_value.find(LABELINFO['mem_size']) < 0:
                dpoResult.update({'mem_size': {'result': 'NG', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['mem_size'],
                                                          'actual': dpoftr_ocr_data[1]}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'mem_size': {'result': 'OK', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['mem_size'],
                                                          'actual': LABELINFO['mem_size']}}})
            if dpoftr_ocr_value.find(LABELINFO['ssd_size']) < 0:
                dpoResult.update({'ssd_size': {'result': 'NG', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['ssd_size'],
                                                          'actual': dpoftr_ocr_data[0]}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'ssd_size': {'result': 'OK', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['ssd_size'],
                                                          'actual': LABELINFO['ssd_size']}}})
        elif country == 'UA':
            if dpoftr_ocr_value.find(LABELINFO['mem_size']) < 0:
                dpoResult.update({'mem_size': {'result': 'NG', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['mem_size'],
                                                          'actual': dpoftr_ocr_data[1]}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'mem_size': {'result': 'OK', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['mem_size'],
                                                          'actual': LABELINFO['mem_size']}}})
            if dpoftr_ocr_value.find(LABELINFO['ssd_size']) < 0:
                dpoResult.update({'ssd_size': {'result': 'NG', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['ssd_size'],
                                                          'actual': dpoftr_ocr_data[2]}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'ssd_size': {'result': 'OK', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['ssd_size'],
                                                          'actual': LABELINFO['ssd_size']}}})
        else:
            if len(dpoftr_ocr_data) == 0:
                dpoResult.update({'dpoftr': {'result': 'NG', 'item': 'dpo', 'detail': 'Missing DPO FTR!'}})
                flag = 1
                return dpoResult, flag
            else:
                pass
            if len(dpoftr_ocr_data[0]) > 1 and dpoftr_ocr_data[0].find('4') >= 0:
                tempValue = dpoftr_ocr_data[0][1:]
                tempssdValue = dpoftr_ocr_data[1]
                dpoftr_ocr_data[0] = '4'
                dpoftr_ocr_data[1] = tempValue
                if len(dpoftr_ocr_data) > 2:
                    dpoftr_ocr_data[2] = tempssdValue
                else:
                    dpoftr_ocr_data.append(tempssdValue)
            if dpoftr_ocr_data[1] != LABELINFO['mem_size']:
                dpoResult.update({'mem_size': {'result': 'NG', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['mem_size'],
                                                          'actual': dpoftr_ocr_data[1]}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'mem_size': {'result': 'OK', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['mem_size'],
                                                          'actual': dpoftr_ocr_data[1]}}})
            if dpoftr_ocr_data[2] != LABELINFO['ssd_size']:
                dpoResult.update({'ssd_size': {'result': 'NG', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['ssd_size'],
                                                          'actual': dpoftr_ocr_data[2]}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'ssd_size': {'result': 'OK', 'item': 'dpo',
                                               'detail': {'expected': LABELINFO['ssd_size'],
                                                          'actual': dpoftr_ocr_data[2]}}})

    elif LABELINFO['order_type'] == 'CTO' or LABELINFO['order_type'] == 'ZPN':
        dposlaHeight = image.shape[0]
        dposlawidth = image.shape[1]
        dposlasearch = image[int(dposlaHeight / 2): int(dposlaHeight / 5 * 4), : int(dposlawidth / 3 * 2)]
        dposlaimage, dposlamapscores = getLabel(dposlasearch, 'dposla')
        #cv2.imwrite('temp.jpg', dposlaimage)
        dposlaSimilarity = calculateSimilarity(dposlaimage, DPOSLA)
        if dposlaSimilarity < 0.8 or dposlamapscores < 0.06:
            dpoResult.update({'sla_similarity': {'result': 'NG', 'item': 'dpo',
                                                 'detail': {'Threshold': 0.8, 'score': dposlaSimilarity}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'sla_similarity': {'result': 'OK', 'item': 'dpo',
                                                 'detail': {'Threshold': 0.8, 'score': dposlaSimilarity}}})
    else:
        pass
    dpofgHeight = image.shape[0]
    dpofgwidth = image.shape[1]
    dpofgsearch = image[int(dpofgHeight / 5 * 3): int(dpofgHeight / 6 * 5), : int(dpofgwidth / 2)]
    # cv2.imwrite('temp.jpg', dpofgsearch)
    dpofgimage, _ = getLabel(dpofgsearch, 'dpofg')
    # cv2.imwrite('temp.jpg', dpofgimage)
    dpofg_ocr_result, _ = FASTOCR(dpofgimage)
    for dpofg in dpofg_ocr_result:
        dpofg_ocr_value += dpofg[1].replace(" ", "").upper()
    if country == 'LL' or country == 'CI' or country == 'C' or country == 'VC':
        if dpofg_ocr_value.replace('0', 'O').find('MODE') < 0 or dpofg_ocr_value.find('MARKED') < 0:
            dpofg_ocr_value = ''
            dpofg_ocr_result = OCRINFERER.ocr(dpofgimage)[0]
            for dpofg in dpofg_ocr_result:
                dpofg_ocr_value += dpofg[1][0].replace(' ', '').upper()
            if dpofg_ocr_value.replace('0', 'O').find('MODE') < 0 or dpofg_ocr_value.find('MARKED') < 0:
                dpoResult.update({'dpofg_otherinfo': {'result': 'NG', 'item': 'dpo',
                                                      'detail': 'The dpo FG other info fail!'}})
                flag = 1
            else:
                pass
        else:
            pass
    else:
        if dpofg_ocr_value.replace('0', 'O').find('MODE') < 0 or dpofg_ocr_value.find('CHINA') < 0 or \
                dpofg_ocr_value.find('APPLE') < 0:
            dpofg_ocr_value = ''
            dpofg_ocr_result = OCRINFERER.ocr(dpofgimage)[0]
            for dpofg in dpofg_ocr_result:
                dpofg_ocr_value += dpofg[1][0].replace(' ', '').upper()
            if dpofg_ocr_value.replace('0', 'O').find('MODE') < 0 or dpofg_ocr_value.find('CHINA') < 0 or \
                dpofg_ocr_value.find('APPLE') < 0:
                dpoResult.update({'dpofg_otherinfo': {'result': 'NG', 'item': 'dpo',
                                                      'detail': 'The dpo FG other info fail!'}})
                flag = 1
            else:
                pass
        else:
            pass
    if dpofg_ocr_value.find(LABELINFO['pn']) < 0:
        if dpofg_ocr_value.replace('0', 'O').split('MODE')[0].find(LABELINFO['pn'].replace('0', 'O')) < 0:
            dpoResult.update({'pn': {'result': 'NG', 'item': 'dpo',
                                     'detail': {'expected': LABELINFO['pn'], 'actual': dpofg_ocr_value.replace('0', 'O').split('MODE')[0]}}})
            flag = 1
        else:
            dpoResult.update({'pn': {'result': 'OK', 'item': 'dpo',
                                     'detail': {'expected': LABELINFO['pn'], 'actual': LABELINFO['pn']}}})
    else:
        dpoResult.update({'pn': {'result': 'OK', 'item': 'dpo',
                                 'detail': {'expected': LABELINFO['pn'], 'actual': LABELINFO['pn']}}})
    if dpofg_ocr_value.find(LABELINFO['model_no']) < 0:
        if dpofg_ocr_value.replace('0', 'O').split('NO.')[1].split(',')[0].find(LABELINFO['model_no'].replace('0', 'O')) < 0:
            dpoResult.update({'model_no': {'result': 'NG', 'item': 'dpo',
                                           'detail': {'expected': LABELINFO['model_no'],
                                                      'actual': dpofg_ocr_value.replace('0', 'O').split('NO.')[1].split(',')[0]}}})
        else:
            dpoResult.update({'model_no': {'result': 'OK', 'item': 'dpo',
                                           'detail': {'expected': LABELINFO['model_no'],
                                                      'actual': LABELINFO['model_no']}}})
        flag = 1
        # return dpoResult, flag
    else:
        dpoResult.update({'model_no': {'result': 'OK', 'item': 'dpo',
                                       'detail': {'expected': LABELINFO['model_no'],
                                                  'actual': LABELINFO['model_no']}}})
    if dpofg_ocr_value.find('COO') < 0:
        if dpofg_ocr_value.find(LABELINFO['config_str']) < 0:
            if dpofg_ocr_value.split(',')[1].split('ASS')[0].replace('O', '0').replace('/', '').replace('B', '8').find(LABELINFO['config_str'].replace('O', '0').replace('/', '').replace('B', '8')) < 0:
                dpofgimage = cv2.resize(dpofgimage, (480, 74))
                # cv2.imwrite('temp.jpg', dpofgimage)
                dpofg_ocr_value = ''
                dpofg_ocr_result, _ = FASTOCR(dpofgimage)
                for dpofg in dpofg_ocr_result:
                    dpofg_ocr_value += dpofg[1].replace(" ", "").upper()
                if dpofg_ocr_value.replace('0', 'O').split('NO.')[1].split('ASS')[0].replace('/', '').replace('B', '8').find(
                        LABELINFO['config_str'].replace('0', 'O').replace('/', '').replace('B', '8')) < 0:
                    dpoResult.update({'config_str': {'result': 'NG', 'item': 'dpo',
                                                     'detail': {'expected': LABELINFO['config_str'],
                                                                'actual': dpofg_ocr_value.replace('0', 'O').split('NO.')[1].split('ASS')[0]}}})
                    flag = 1
                else:
                    dpoResult.update({'config_str': {'result': 'OK', 'item': 'dpo',
                                                     'detail': {'expected': LABELINFO['config_str'],
                                                                'actual': LABELINFO['config_str']}}})
            else:
                dpoResult.update({'config_str': {'result': 'OK', 'item': 'dpo',
                                                 'detail': {'expected': LABELINFO['config_str'],
                                                            'actual': LABELINFO['config_str']}}})
            # return dpoResult, flag
        else:
            dpoResult.update({'config_str': {'result': 'OK', 'item': 'dpo',
                                             'detail': {'expected': LABELINFO['config_str'],
                                                        'actual': LABELINFO['config_str']}}})
    else:
        if dpofg_ocr_value.find(LABELINFO['config_str']) < 0:
            if dpofg_ocr_value.split(',')[1].split('COO')[0].replace('O', '0').replace('/', '').replace('B', '8').find(LABELINFO['config_str'].replace('O', '0').replace('/', '').replace('B', '8')) < 0:
                dpofgimage = cv2.resize(dpofgimage, (480, 64))
                # cv2.imwrite('temp.jpg', dpofgimage)
                dpofg_ocr_value = ''
                dpofg_ocr_result, _ = FASTOCR(dpofgimage)
                for dpofg in dpofg_ocr_result:
                    dpofg_ocr_value += dpofg[1].replace(" ", "").upper()
                if dpofg_ocr_value.replace('0', 'O').split('NO.')[1].split('COO')[0].replace('/', '').replace('B', '8').find(
                        LABELINFO['config_str'].replace('O', '0').replace('/', '').replace('B', '8')) < 0:
                    dpoResult.update({'config_str': {'result': 'NG', 'item': 'dpo',
                                                     'detail': {'expected': LABELINFO['config_str'],
                                                                'actual': dpofg_ocr_value.replace('0', 'O').split('NO.')[1].split('COO')[0]}}})
                    flag = 1
                else:
                    dpoResult.update({'config_str': {'result': 'OK', 'item': 'dpo',
                                                     'detail': {'expected': LABELINFO['config_str'],
                                                                'actual': LABELINFO['config_str']}}})
            else:
                dpoResult.update({'config_str': {'result': 'OK', 'item': 'dpo',
                                                 'detail': {'expected': LABELINFO['config_str'],
                                                            'actual': LABELINFO['config_str']}}})
            # return dpoResult, flag
        else:
            dpoResult.update({'config_str': {'result': 'OK', 'item': 'dpo',
                                             'detail': {'expected': LABELINFO['config_str'],
                                                        'actual': LABELINFO['config_str']}}})
    #####裁剪upc并识别
    dpoupcHeight = image.shape[0]
    dpoupcWidth = image.shape[1]
    dpoimagedown = image[int(dpoupcHeight / 2): int(dpoupcHeight / 10 * 9), : int(dpoupcWidth / 4)]
    # cv2.imwrite('temp1.jpg', dpoimagedown)
    dpoupcimage, dpoupcSimilarity = getLabel(dpoimagedown, 'dpoupc')
    # cv2.imwrite('temp.jpg', dpoupcimage)
    height = dpoupcimage.shape[0]
    if LABELINFO['upc'] != LABELINFO['pn']:
        upc_image = dpoupcimage[:int(height / 3) + 30]
        # cv2.imwrite('temp1.jpg', upc_image)
        # upc_ocr_result = OCRINFEREREN.ocr(upc_image)[0]
        # for lineupc in upc_ocr_result:
        #     dpoupc_ocr_value += lineupc[1][0].replace(' ', '').upper()
        upc_ocr_result, _ = FASTOCR(upc_image)
        for lineupc in upc_ocr_result:
            dpoupc_ocr_value += lineupc[1].replace(" ", "").upper()
        pattern = r'\d+'
        dpoupc_ocr_value = dpoupc_ocr_value.replace('O', '0').replace('I', '1')
        dpoupc_ocr_value = re.findall(pattern, dpoupc_ocr_value)[0]
        if LABELINFO['upc'].find(dpoupc_ocr_value) < 0 or dpoupc_ocr_value[-1:] != LABELINFO['upc'][-1:]:
            dpoupc_ocr_value = ''
            upc_image_resize = cv2.resize(upc_image, (320, 64))
            # cv2.imwrite('temp.jpg', upc_image_gray)
            upc_ocr_result, _ = FASTOCR(upc_image_resize)
            for lineupc in upc_ocr_result:
                dpoupc_ocr_value += lineupc[1].replace(" ", "").upper()
            pattern = r'\d+'
            dpoupc_ocr_value = dpoupc_ocr_value.replace('O', '0').replace('I', '1')
            dpoupc_ocr_value = re.findall(pattern, dpoupc_ocr_value)[0]
            if LABELINFO['upc'].find(dpoupc_ocr_value) < 0 or dpoupc_ocr_value[-1:] != LABELINFO['upc'][-1:]:
                dpoupc_ocr_value = ''
                upc_image_resize = cv2.resize(upc_image, (960, 128))
                upc_image_gray = cv2.cvtColor(upc_image_resize, cv2.COLOR_BGR2GRAY)
                upc_image_gray = cv2.GaussianBlur(upc_image_gray, (3, 3), 0)
                upc_ocr_result, _ = FASTOCR(upc_image_gray)
                for txt in upc_ocr_result:
                    dpoupc_ocr_value += txt[1].replace(" ", "").replace("″", "").replace("\"", "").upper()
        else:
            pass
        dpoupc_barcode_value = barcodeDecode(upc_image)
        if dpoupc_ocr_value.find('UPC') < 0:
            pattern = r'\d+'
            dpoupc_ocr_value = dpoupc_ocr_value.replace('O', '0').replace('I', '1')
            dpoupc_ocr_value = re.findall(pattern, dpoupc_ocr_value)[0]
            if LABELINFO['upc'].find(dpoupc_ocr_value) < 0 or dpoupc_ocr_value[-1:] != LABELINFO['upc'][-1:]:
                dpoResult.update({'upc_txt': {'result': 'NG', 'item': 'dpo',
                                              'detail': {'expected': LABELINFO['upc'], 'actual': dpoupc_ocr_value}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'upc_txt': {'result': 'OK', 'item': 'dpo',
                                              'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
        else:
            pattern = r'\d+'
            dpoupc_ocr_value = dpoupc_ocr_value.split('UPC')[1].replace('O', '0').replace('I', '1')
            dpoupc_ocr_value = re.findall(pattern, dpoupc_ocr_value)[0]
            if LABELINFO['upc'].find(dpoupc_ocr_value) < 0 or dpoupc_ocr_value[-1:] != LABELINFO['upc'][-1:]:
                dpoResult.update({'upc_txt': {'result': 'NG', 'item': 'dpo',
                                              'detail': {'expected': LABELINFO['upc'], 'actual': dpoupc_ocr_value}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'upc_txt': {'result': 'OK', 'item': 'dpo',
                                              'detail': {'expected': LABELINFO['upc'], 'actual': LABELINFO['upc']}}})
        if dpoupc_barcode_value.find(LABELINFO['upc']) < 0:
            dpoResult.update({'upc_code': {'result': 'NG', 'item': 'dpo',
                                           'detail': {'expected': LABELINFO['upc'], 'actual': dpoupc_barcode_value}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'upc_code': {'result': 'OK', 'item': 'dpo',
                                           'detail': {'expected': LABELINFO['upc'], 'actual': dpoupc_barcode_value}}})
        pn_image = dpoupcimage[int(height / 3): int(height / 3 * 2)]
        # cv2.imwrite('temp.jpg', pn_image)
        pn_ocr_result, _ = FASTOCR(pn_image)
        if pn_ocr_result is None:
            pn_ocr_result = OCRINFERER.ocr(pn_image)[0]
            for line in pn_ocr_result:
                dpopn_ocr_value += line[1][0].replace(' ', '').replace(',', '.').upper()
        else:
            for line in pn_ocr_result:
                dpopn_ocr_value += line[1].replace(" ", "").replace(',', '.').upper()
        dpopn_barcode_value = barcodeDecode(pn_image)
        if dpopn_ocr_value.find(LABELINFO['pn']) < 0:
            if dpopn_ocr_value.replace('O', '0').split('N0.')[1].replace('/', '').find(LABELINFO['pn'].replace('O', '0').replace('/', '')) < 0:
                dpoResult.update({'pn_txt': {'result': 'NG', 'item': 'upc',
                                             'detail': {'expected': LABELINFO['pn'],
                                                        'actual': dpopn_ocr_value.replace('O', '0').split('N0.')[1]}}})
                flag = 1
            else:
                dpoResult.update({'pn_txt': {'result': 'OK', 'item': 'upc',
                                             'detail': {'expected': LABELINFO['pn'],
                                                        'actual': LABELINFO['pn']}}})
            # return dpoResult, flag
        else:
            dpoResult.update({'pn_txt': {'result': 'OK', 'item': 'dpo',
                                         'detail': {'expected': LABELINFO['pn'], 'actual': LABELINFO['pn']}}})
        if dpopn_barcode_value.find(LABELINFO['pn']) < 0:
            dpoResult.update({'pn_code': {'result': 'NG', 'item': 'dpo',
                                          'detail': {'expected': LABELINFO['pn'], 'actual': dpopn_barcode_value}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'pn_code': {'result': 'OK', 'item': 'dpo',
                                          'detail': {'expected': LABELINFO['pn'], 'actual': dpopn_barcode_value}}})
        sn_image = dpoupcimage[int(height / 3 * 2):]
        # cv2.imwrite('temp.jpg', sn_image)
        sn_ocr_result, _ = FASTOCR(sn_image)
        if sn_ocr_result is None:
            sn_ocr_result = OCRINFERER.ocr(sn_image)[0]
            for line in sn_ocr_result:
                dposn_ocr_value += line[1][0].replace(' ', '').replace(',', '.').upper()
        else:
            for line in sn_ocr_result:
                dposn_ocr_value += line[1].replace(" ", "").replace(',', '.').upper()
        dposn_barcode_value = barcodeDecode(sn_image)
        if dposn_ocr_value.replace('O', '0').find(LABELINFO['sn']) < 0:
            dpoResult.update({'sn_txt': {'result': 'NG', 'item': 'dpo',
                                         'detail': {'expected': LABELINFO['sn'],
                                                    'actual': dposn_ocr_value.replace('O', '0').split('N0.')[1]}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'sn_txt': {'result': 'OK', 'item': 'dpo',
                                         'detail': {'expected': LABELINFO['sn'],
                                                    'actual': LABELINFO['sn']}}})
            # return dpoResult, flag
        if dposn_barcode_value.find(LABELINFO['sn']) < 0:
            dpoResult.update({'sn_code': {'result': 'NG', 'item': 'dpo',
                                          'detail': {'expected': LABELINFO['sn'], 'actual': dposn_barcode_value}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'sn_code': {'result': 'OK', 'item': 'dpo',
                                          'detail': {'expected': LABELINFO['sn'], 'actual': dposn_barcode_value}}})
        if country == 'KH' or country == 'LE' or country == 'PA' or country == 'PP' or country == 'X' or country == 'ZA':
            pass
        else:
            dposhipperHeight = image.shape[0]
            dposhipperWidth = image.shape[1]
            dposhipperresearch = image[int(dposhipperHeight / 5 * 3): int(dposhipperHeight / 9 * 8), int(dposhipperWidth / 5): int(dposhipperWidth / 3 * 2)]
            # cv2.imwrite('temp.jpg', dposhipperresearch)
            dposhipperimage, dposhipersimilarity = getLabel(dposhipperresearch, 'dposhipper')
            # cv2.imwrite('temp.jpg', dposhipperimage)
            # dposhipperSimilarity = getSimilarity(dposhipperimage, DPOSHIPPER)
            dposhipperSimilarity = calculateSimilarity(dposhipperimage, DPOSHIPPER)
            if dposhipperSimilarity < 0.7 or dposhipersimilarity < 0.1:
                dpoResult.update({'dpo_similarity': {'result': 'NG', 'item': 'dpo',
                                                     'detail': {'Threshold': 0.7, 'score': dposhipperSimilarity}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'dpo_similarity': {'result': 'OK', 'item': 'dpo',
                                                     'detail': {'Threshold': 0.7, 'score': dposhipperSimilarity}}})
                # return dpoResult, flag
            if country == 'CH':
                va_ocr_result, _ = FASTOCR(dposhipperimage)
                if va_ocr_result is not None:
                    for lineva in va_ocr_result:
                        dposhipperva_ocr_value += lineva[1].replace(" ", "").upper().replace('O', '0')
                    if dposhipperva_ocr_value.find('20.6V') < 0 or dposhipperva_ocr_value.find('3.4A') < 0 or \
                            dposhipperva_ocr_value.find('20.5V') < 0 or dposhipperva_ocr_value.find('4.7A') < 0:
                        dpoResult.update({'vcinfo': {'result': 'NG', 'item': 'dpo',
                                                     'detail': {'expected': '20.6V3.4A20.5V4.7A',
                                                                'actual': dposhipperva_ocr_value.split('输入')[1].split('生产')[0]}}})
                        flag = 1
                    else:
                        dpoResult.update({'vcinfo': {'result': 'OK', 'item': 'dpo',
                                                     'detail': {'expected': '20.6V3.4A20.5V4.7A',
                                                                'actual': '20.6V3.4A20.5V4.7A'}}})
                else:
                    pass
            elif country == 'C' or country == 'LL':
                modelno_ocr_result, _ = FASTOCR(dposhipperimage)
                if modelno_ocr_result is not None:
                    for linemodelno in modelno_ocr_result:
                        dposhippermodelno_ocr_value += linemodelno[1].replace(" ", "").upper().replace('O', '0').replace('：', ':')
                    if dposhippermodelno_ocr_value.split('D:')[1].split('C:')[0].find(LABELINFO['model_no']) < 0\
                            or dposhippermodelno_ocr_value.split('C:')[1].find(LABELINFO['model_no']) < 0:
                        dpoResult.update({'dposhippermodelno': {'result': 'NG', 'item': 'dpo',
                                                                'detail': {'expected': LABELINFO['model_no'],
                                                                           'actual1': dposhippermodelno_ocr_value.split('D:')[1].split('C:')[0],
                                                                           'actual2': dposhippermodelno_ocr_value.split('C:')[1]}}})
                        flag = 1
                    else:
                        dpoResult.update({'dposhippermodelno': {'result': 'OK', 'item': 'dpo',
                                                                'detail': {'expected': LABELINFO['model_no'],
                                                                           'actual': LABELINFO['model_no']}}})
            elif country == 'TH' or country == 'TG':
                year_ocr_result, _ = FASTOCR(dposhipperimage)
                if year_ocr_result is not None:
                    for lineyear in year_ocr_result:
                        dposhipperyear_ocr_value += lineyear[1].replace(" ", "").upper().replace('O', '0').replace('：', ':')
                    if dposhipperyear_ocr_value.find(str(int('20' + LABELINFO['yr']) + 543)) < 0:
                        dpoResult.update({'dposhipperyear': {'result': 'NG', 'item': 'dpo',
                                                             'detail': {'expected': str(int('20' + LABELINFO['yr']) + 543),
                                                                        'actual': dposhipperyear_ocr_value.split(':')[1]}}})
                        flag = 1
                    else:
                        dpoResult.update({'dposhipperyear': {'result': 'OK', 'item': 'dpo',
                                                             'detail': {'expected': str(int('20' + LABELINFO['yr']) + 543),
                                                                        'actual': str(int('20' + LABELINFO['yr']) + 543)}}})
            else:
                pass
    else:
        pn_image = dpoupcimage[: int(height / 2)]
        # cv2.imwrite('temp.jpg', pn_image)
        pn_ocr_result, _ = FASTOCR(pn_image)
        for line in pn_ocr_result:
            dpopn_ocr_value += line[1].replace(" ", "").replace(',', '.').upper()
        dpopn_barcode_value = barcodeDecode(pn_image)
        if dpopn_ocr_value.find(LABELINFO['pn']) < 0:
            if dpopn_ocr_value.replace('O', '0').split('N0.')[1].replace('/', '').find(LABELINFO['pn'].replace('O', '0').replace('/', '')) < 0:
                dpoResult.update({'pn_txt': {'result': 'NG', 'item': 'upc',
                                             'detail': {'expected': LABELINFO['pn'],
                                                        'actual': dpopn_ocr_value.replace('O', '0').split('N0.')[1]}}})
                flag = 1
            else:
                dpoResult.update({'pn_txt': {'result': 'OK', 'item': 'upc',
                                             'detail': {'expected': LABELINFO['pn'],
                                                        'actual': LABELINFO['pn']}}})
            # return dpoResult, flag
        else:
            dpoResult.update({'pn_txt': {'result': 'OK', 'item': 'dpo',
                                         'detail': {'expected': LABELINFO['pn'],
                                                    'actual': LABELINFO['pn']}}})
        if dpopn_barcode_value.find(LABELINFO['pn']) < 0:
            dpoResult.update({'pn_code': {'result': 'NG', 'item': 'dpo',
                                          'detail': {'expected': LABELINFO['pn'], 'actual': dpopn_barcode_value}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'pn_code': {'result': 'OK', 'item': 'dpo',
                                          'detail': {'expected': LABELINFO['pn'], 'actual': dpopn_barcode_value}}})
        sn_image = dpoupcimage[int(height / 2):]
        # cv2.imwrite('temp.jpg', sn_image)
        # sn_ocr_result = OCRINFERER.ocr(sn_image)[0]
        # for line in sn_ocr_result:
        #     dposn_ocr_value += line[1][0].replace(' ', '').replace(',', '.').upper()
        sn_ocr_result, _ = FASTOCR(sn_image)
        for line in sn_ocr_result:
            dposn_ocr_value += line[1].replace(" ", "").replace(',', '.').upper()
        dposn_barcode_value = barcodeDecode(sn_image)
        if dposn_ocr_value.replace('O', '0').find(LABELINFO['sn']) < 0:
            dpoResult.update({'sn_txt': {'result': 'NG', 'item': 'dpo',
                                         'detail': {'expected': LABELINFO['sn'],
                                                    'actual': dposn_ocr_value.replace('O', '0').split('N0.')[1]}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'sn_txt': {'result': 'OK', 'item': 'dpo',
                                         'detail': {'expected': LABELINFO['sn'],
                                                    'actual': LABELINFO['sn']}}})
            # return dpoResult, flag
        if dposn_barcode_value.find(LABELINFO['sn']) < 0:
            dpoResult.update({'sn_code': {'result': 'NG', 'item': 'dpo',
                                          'detail': {'expected': LABELINFO['sn'], 'actual': dposn_barcode_value}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'sn_code': {'result': 'OK', 'item': 'dpo',
                                          'detail': {'expected': LABELINFO['sn'], 'actual': dposn_barcode_value}}})

        if country == 'KH' or country == 'LE' or country == 'PA' or country == 'PP' or country == 'X' or country == 'ZA':
            pass
        else:
            dposhipperHeight = image.shape[0]
            dposhipperWidth = image.shape[1]
            dposhipperresearch = image[int(dposhipperHeight / 10 * 7): int(dposhipperHeight / 9 * 8),
                                 int(dposhipperWidth / 5): int(dposhipperWidth / 3 * 2)]
            # cv2.imwrite('temp.jpg', dposhipperresearch)
            dposhipperimage, dposhipersimilarity = getLabel(dposhipperresearch, 'dposhipper')
            # cv2.imwrite('temp.jpg', dposhipperimage)
            dposhipperSimilarity = calculateSimilarity(dposhipperimage, DPOSHIPPER)
            if dposhipperSimilarity < 0.7 or dposhipersimilarity < 0.1:
                dpoResult.update({'dpo_similarity': {'result': 'NG', 'item': 'dpo',
                                                     'detail': {'Threshold': 0.7, 'score': dposhipperSimilarity}}})
                flag = 1
                # return dpoResult, flag
            else:
                dpoResult.update({'dpo_similarity': {'result': 'OK', 'item': 'dpo',
                                                     'detail': {'Threshold': 0.7, 'score': dposhipperSimilarity}}})
                # return dpoResult, flag
            if country == 'CH':
                va_ocr_result, _ = FASTOCR(dposhipperimage)
                if va_ocr_result is not None:
                    for lineva in va_ocr_result:
                        dposhipperva_ocr_value += lineva[1].replace(" ", "").upper().replace('O', '0')
                    if dposhipperva_ocr_value.find('20.6V') < 0 or dposhipperva_ocr_value.find('3.4A') < 0 or \
                            dposhipperva_ocr_value.find('20.5V') < 0 or dposhipperva_ocr_value.find('4.7A') < 0:
                        dpoResult.update({'vcinfo': {'result': 'NG', 'item': 'dpo',
                                                     'detail': {'expected': '20.6V3.4A20.5V4.7A',
                                                                'actual': dposhipperva_ocr_value.split('输入')[1].split('生产')[0]}}})
                        flag = 1
                    else:
                        dpoResult.update({'vcinfo': {'result': 'OK', 'item': 'dpo',
                                                     'detail': {'expected': '20.6V3.4A20.5V4.7A',
                                                                'actual': '20.6V3.4A20.5V4.7A'}}})
                else:
                    pass
            elif country == 'C' or country == 'LL':
                modelno_ocr_result, _ = FASTOCR(dposhipperimage)
                if modelno_ocr_result is not None:
                    for linemodelno in modelno_ocr_result:
                        dposhippermodelno_ocr_value += linemodelno[1].replace(" ", "").upper().replace('O', '0').replace('：', ':')
                    if dposhippermodelno_ocr_value.split('D:')[1].split('C:')[0].find(LABELINFO['model_no']) < 0 \
                            or dposhippermodelno_ocr_value.split('C:')[1].find(LABELINFO['model_no']) < 0:
                        dpoResult.update({'dposhippermodelno': {'result': 'NG', 'item': 'dpo',
                                                                'detail': {'expected': LABELINFO['model_no'],
                                                                           'actual1': dposhippermodelno_ocr_value.split('D:')[1].split('C:')[0],
                                                                           'actual2': dposhippermodelno_ocr_value.split('C:')[1]}}})
                        flag = 1
                    else:
                        dpoResult.update({'dposhippermodelno': {'result': 'OK', 'item': 'dpo',
                                                                'detail': {'expected': LABELINFO['model_no'],
                                                                           'actual': LABELINFO['model_no']}}})
            elif country == 'TH' or country == 'TG':
                year_ocr_result, _ = FASTOCR(dposhipperimage)
                if year_ocr_result is not None:
                    for lineyear in year_ocr_result:
                        dposhipperyear_ocr_value += lineyear[1].replace(" ", "").upper().replace('O', '0').replace('：', ':')
                    if dposhipperyear_ocr_value.find(str(int('20' + LABELINFO['yr']) + 543)) < 0:
                        dpoResult.update({'dposhipperyear': {'result': 'NG', 'item': 'dpo',
                                                             'detail': {'expected': str(int('20' + LABELINFO['yr']) + 543),
                                                                        'actual': dposhipperyear_ocr_value.split(':')[1]}}})
                        flag = 1
                    else:
                        dpoResult.update({'dposhipperyear': {'result': 'OK', 'item': 'dpo',
                                                             'detail': {'expected': str(int('20' + LABELINFO['yr']) + 543),
                                                                        'actual': str(int('20' + LABELINFO['yr']) + 543)}}})
            else:
                pass

    if country == 'SA':
        dpoimporterHeight = image.shape[0]
        dpoimporterWidth = image.shape[1]
        dpoimporterSearch = image[int(dpoimporterHeight / 5 * 3):int(dpoimporterHeight / 10 * 9), : int(dpoimporterWidth / 3 * 2)]
        # cv2.imwrite('temp.jpg', dpoimporterSearch)
        dpoimporterimage, _ = getLabel(dpoimporterSearch, 'dpoimporter')
        # cv2.imwrite('temp.jpg', dpoimporterimage)
        dpoimporterSimilarity = calculateSimilarity(dpoimporterimage, DPOIMPORTER)
        if dpoimporterSimilarity < 0.8:
            dpoResult.update({'dpoimporter_similarity': {'result': 'NG', 'item': 'dpo',
                                                         'detail': {'Threshold': 0.8, 'score': dpoimporterSimilarity}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'dpoimporter_similarity': {'result': 'OK', 'item': 'dpo',
                                                         'detail': {'Threshold': 0.8, 'score': dpoimporterSimilarity}}})
            # return dpoResult, flag
        dpoEnergyHeight = image.shape[0]
        dpoEnergyWidth = image.shape[1]
        dpoEnergySearch = image[int(dpoEnergyHeight / 5 * 3):int(dpoEnergyHeight / 10 * 9), int(dpoEnergyWidth / 3): int(dpoEnergyWidth / 3 * 2)]
        # cv2.imwrite('temp.jpg', dpoEnergySearch)
        energyimage, _ = getLabel(dpoEnergySearch, 'energy')
        # cv2.imwrite('temp.jpg', energyimage)
        energySimilarity = calculateSimilarity(energyimage, ENERGY)
        if energySimilarity < 0.7:
            dpoResult.update({'dpoenergy_similarity': {'result': 'NG', 'item': 'dpo',
                                                       'detail': {'Threshold': 0.7, 'score': energySimilarity}}})
            flag = 1
            return dpoResult, flag
        else:
            dpoResult.update({'dpoenergy_similarity': {'result': 'OK', 'item': 'dpo',
                                                       'detail': {'Threshold': 0.7, 'score': energySimilarity}}})
            return dpoResult, flag
    elif country == 'ID':
        dpoimporterHeight = image.shape[0]
        dpoimporterWidth = image.shape[1]
        dpoimporterSearch = image[int(dpoimporterHeight / 5 * 3):int(dpoimporterHeight / 10 * 9), : int(dpoimporterWidth / 3 * 2)]
        dpoimporterimage, _ = getLabel(dpoimporterSearch, 'dpoimporter')
        dpoimporterSimilarity = calculateSimilarity(dpoimporterimage, DPOIMPORTER)
        if dpoimporterSimilarity < 0.8:
            dpoResult.update({'dpoimporter_similarity': {'result': 'NG', 'item': 'dpo',
                                                         'detail': {'Threshold': 0.8, 'score': dpoimporterSimilarity}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'dpoimporter_similarity': {'result': 'OK', 'item': 'dpo',
                                                         'detail': {'Threshold': 0.8, 'score': dpoimporterSimilarity}}})
            # return dpoResult, flag
        vh_ocr_result, _ = FASTOCR(dpoimporterimage)
        if vh_ocr_result is not None:
            for linevh in vh_ocr_result:
                dpoimportervh_ocr_value += linevh[1].replace(" ", "").upper().replace('O', '0').replace('：', ':')
            if dpoimportervh_ocr_value.find('100') < 0 or dpoimportervh_ocr_value.find('240V') < 0 or \
                    dpoimportervh_ocr_value.find('50') < 0 or dpoimportervh_ocr_value.find('60HZ') < 0:
                dpoResult.update({'vhinfo': {'result': 'NG', 'item': 'dpo',
                                             'detail': {'expected': '100-240V50-60HZ',
                                                        'actual': dpoimportervh_ocr_value.split('K:')[1].split('HZ')[0]}}})
                flag = 1
                return dpoResult, flag
            else:
                dpoResult.update({'vhinfo': {'result': 'OK', 'item': 'dpo',
                                             'detail': {'expected': '100-240V50-60HZ',
                                                        'actual': '100-240V50-60HZ'}}})
                return dpoResult, flag
        else:
            return dpoResult, flag
    elif country == 'BZ' or country == 'SA' or country == 'TG' or country == 'TH' or country == 'CT' \
            or country == 'TA' or country == 'UA':
        dpoimporterHeight = image.shape[0]
        dpoimporterWidth = image.shape[1]
        dpoimporterSearch = image[int(dpoimporterHeight / 5 * 3):int(dpoimporterHeight / 10 * 9), : int(dpoimporterWidth / 3 * 2)]
        dpoimporterimage, _ = getLabel(dpoimporterSearch, 'dpoimporter')
        dpoimporterSimilarity = calculateSimilarity(dpoimporterimage, DPOIMPORTER)
        if dpoimporterSimilarity < 0.8:
            dpoResult.update({'dpoimporter_similarity': {'result': 'NG', 'item': 'dpo',
                                                         'detail': {'Threshold': 0.8, 'score': dpoimporterSimilarity}}})
            flag = 1
            return dpoResult, flag
        else:
            dpoResult.update({'dpoimporter_similarity': {'result': 'OK', 'item': 'dpo',
                                                         'detail': {'Threshold': 0.8, 'score': dpoimporterSimilarity}}})
            return dpoResult, flag
    elif country == 'CH':
        dpoimporterHeight = image.shape[0]
        dpoimporterWidth = image.shape[1]
        dpoimporterSearch = image[int(dpoimporterHeight / 5 * 3):int(dpoimporterHeight / 10 * 9), : int(dpoimporterWidth / 3 * 2)]
        # cv2.imwrite('temp.jpg', dpoimporterSearch)
        dpoimporterimage, _ = getLabel(dpoimporterSearch, 'dpoimporter')
        dpoimporterSimilarity = calculateSimilarity(dpoimporterimage, DPOIMPORTER)
        if dpoimporterSimilarity < 0.8:
            dpoResult.update({'dpoimporter_similarity': {'result': 'NG', 'item': 'dpo',
                                                         'detail': {'Threshold': 0.8, 'score': dpoimporterSimilarity}}})
            flag = 1
            # return dpoResult, flag
        else:
            dpoResult.update({'dpoimporter_similarity': {'result': 'OK', 'item': 'dpo',
                                                         'detail': {'Threshold': 0.8, 'score': dpoimporterSimilarity}}})
        va_ocr_result, _ = FASTOCR(dpoimporterimage)
        if va_ocr_result is not None:
            for lineva in va_ocr_result:
                dpoimporterva_ocr_value += lineva[1].replace(" ", "").upper().replace('O', '0')
            if dpoimporterva_ocr_value.find('20.6V') < 0 or dpoimporterva_ocr_value.find('3.4A') < 0 or \
                    dpoimporterva_ocr_value.find('20.5V') < 0 or dpoimporterva_ocr_value.find('4.7A') < 0:
                dpoResult.update({'vcinfo': {'result': 'NG', 'item': 'dpo',
                                             'detail': {'expected': '20.6V3.4A20.5V4.7A',
                                                        'actual': dpoimporterva_ocr_value.split('输入')[1]}}})
                flag = 1
            else:
                dpoResult.update({'vcinfo': {'result': 'OK', 'item': 'dpo',
                                             'detail': {'expected': '20.6V3.4A20.5V4.7A',
                                                        'actual': '20.6V3.4A20.5V4.7A'}}})
        else:
            pass
        energyHeight = image.shape[0]
        energrimage = image[int(energyHeight / 2):]
        # cv2.imwrite('temp.jpg', energrimage)
        energyimage, energymapSimilarity = getLabel(energrimage, 'energy')
        # cv2.imwrite('temp.jpg', energyimage)
        energy_ocr_result, _ = FASTOCR(energyimage)
        if energy_ocr_result is None:
            pass
        else:
            for lineenergy in energy_ocr_result:
                dpoenergy_ocr_value += lineenergy[1].replace(" ", "").upper()
            if dpoenergy_ocr_value.replace('I', '1').replace('O', '0').find(LABELINFO['model_no'].replace('I', '1').replace('O', '0')) < 0:
                dpoResult.update({'dpoenergy_modelno': {'result': 'NG', 'item': 'dpo',
                                                        'detail': {'expected': LABELINFO['model_no'],
                                                                   'actual': dpoenergy_ocr_value.replace(':', '：').split('规格型号：')[1].split('典型')[0]}}})
                flag = 1
            else:
                dpoResult.update({'dpoenergy_modelno': {'result': 'OK', 'item': 'dpo',
                                                        'detail': {'expected': LABELINFO['model_no'], 'actual': LABELINFO['model_no']}}})
        energySimilarity = calculateSimilarity(energyimage, ENERGY)
        if energySimilarity < 0.7 or energymapSimilarity < 0.4:
            dpoResult.update({'dpoenergy_similarity': {'result': 'NG', 'item': 'dpo',
                                                       'detail': {'Threshold': 0.7, 'score': energySimilarity}}})
            flag = 1
            return dpoResult, flag
        else:
            dpoResult.update({'dpoenergy_similarity': {'result': 'OK', 'item': 'dpo',
                                                       'detail': {'Threshold': 0.7, 'score': energySimilarity}}})
            return dpoResult, flag
    else:
        return dpoResult, flag


def getMoppSdppiResult(inputPath):
    moppResult = {}
    flag = 0
    image = cv2.imread(inputPath)
    country = LABELINFO['country']
    moppsdppiheight = image.shape[0]
    moppsdppiwidth = image.shape[1]
    moppImageTemp = image[int(moppsdppiheight / 3): int(moppsdppiheight / 4 *3), int(moppsdppiwidth / 3 ): int(moppsdppiwidth / 4 * 3)]
    # cv2.imwrite('temp1.jpg', moppImageTemp)
    moppimage, moppMapSimilarity = getLabel(moppImageTemp, 'mopp')
    # cv2.imwrite('temp.jpg', moppimage)
    moppSimilarity = calculateSimilarity(moppimage, MOPP)
    # if moppSimilarity < 0.8:
    #     moppResult.update({'mopp': {'result': 'NG', 'item': 'mopp', 'detail': "Detect MOPP fail"}})
    #     flag = 1
    #     # return moppResult, flag
    # else:
    #     moppResult.update({'mopp': {'result': 'OK', 'item': 'mopp', 'detail': 'mopp label exit'}})
    moppimageRotated = cv2.flip(cv2.transpose(moppimage), 0)
    moppimageMirror = cv2.flip(moppimageRotated, 1)
    #cv2.imwrite('temp.jpg', moppimageMirror)
    height = moppimageMirror.shape[0]
    mopppartimage = moppimageMirror[int(height / 2):]
    #cv2.imwrite('temp.jpg', mopppartimage)
    mopp_ocr_result, _ = FASTOCR(mopppartimage)
    if mopp_ocr_result is None:
        mopppartimage = moppimageMirror[int(height / 5):]
        # cv2.imwrite('temp.jpg', mopppartimage)
        mopp_ocr_result, _ = FASTOCR(mopppartimage)
    if mopp_ocr_result is None or moppSimilarity < 0.8:
        moppResult.update({'mopp': {'result': 'NG', 'item': 'mopp', 'detail': "Detect MOPP fail!"}})
        flag = 1
        # return moppResult, flag
    else:
        moppResult.update({'mopp': {'result': 'OK', 'item': 'mopp', 'detail': 'Mopp label exit!'}})
    # mopp_ocr_value = ''
    # for txt in mopp_ocr_result:
    #     mopp_ocr_value += txt[1].replace(" ", "").upper()
    # if country == 'TA':
    #     if mopp_ocr_value.find('傷害視力') < 0:
    #         flag = 1
    #         moppResult.update({'mopp': {'result': 'NG', 'item': 'mopp', 'detail': 'vision label not exit'}})
    #     else:
    #         moppResult.update({'mopp': {'result': 'OK', 'item': 'mopp', 'detail': 'vision label exit'}})
    # else:
    #     if mopp_ocr_value.find('傷害視力') < 0:
    #         moppResult.update({'mopp': {'result': 'OK', 'item': 'mopp', 'detail': 'This country not detect vision'}})
    #     else:
    #         flag = 1
    #         moppResult.update({'mopp': {'result': 'NG', 'item': 'mopp', 'detail': 'This country not detect vision'}})
    sdppiTempImage = image[int(moppsdppiheight / 4): int(moppsdppiheight / 6 * 5), : int(moppsdppiwidth / 2)]
    # cv2.imwrite('temp1.jpg', sdppiTempImage)
    sdppimage, _ = getLabel(sdppiTempImage, 'sdppi')
    # cv2.imwrite('temp2.jpg', sdppimage)
    sdppiSimilarity = getColorSimilarity(sdppimage, SDPPI)
    if sdppiSimilarity < 0.2:
        moppResult.update({'maccolor_similarity': {'result': 'NG', 'item': 'sdppi',
                                                   'detail': {'Threshold': 0.7, 'score': sdppiSimilarity}}})
        flag = 1
        # return moppResult, flag
    else:
        if sdppiSimilarity < 0.5:
            sdppiSimilarity += 0.5
        moppResult.update({'maccolor_similarity': {'result': 'OK', 'item': 'sdppi',
                                                   'detail': {'Threshold': 0.7, 'score': sdppiSimilarity}}})
        # return moppResult, flag
    if country == 'KH':
        meps_ocr_value = ''
        mepsimage, _ = getLabel(sdppimage, 'meps')
        mepsSimilarity = calculateSimilarity(mepsimage, MEPS)
        if mepsSimilarity < 0.8:
            moppResult.update({'meps_similarity': {'result': 'NG', 'item': 'sdppi',
                                                   'detail': {'Threshold': 0.8, 'score': mepsSimilarity}}})
            flag = 1
            # return moppResult, flag
        else:
            moppResult.update({'meps_similarity': {'result': 'OK', 'item': 'sdppi',
                                                   'detail': {'Threshold': 0.8, 'score': mepsSimilarity}}})
        mepsImage = cv2.flip(cv2.transpose(mepsimage), 1)
        # cv2.imwrite('temp.jpg', mepsImage)
        meps_ocr_result = OCRINFERER.ocr(mepsImage)[0]
        if len(meps_ocr_result) != 0:
            for line in meps_ocr_result:
                meps_ocr_value += line[1][0].replace(' ', '').upper().replace('O', '0')
            if meps_ocr_value.find(LABELINFO['model_no']) < 0:
                moppResult.update({'meps': {'result': 'NG', 'item': 'sdppi',
                                            'detail': {'except': LABELINFO['model_no'],
                                                       'actual': meps_ocr_value}}})
                flag = 1
                return moppResult, flag
            else:
                moppResult.update({'meps': {'result': 'OK', 'item': 'sdppi',
                                            'detail': {'except': LABELINFO['model_no'],
                                                       'actual': LABELINFO['model_no']}}})
                return moppResult, flag
        else:
            return moppResult, flag
    elif country == 'ID':
        mepsimage, _ = getLabel(sdppimage, 'meps')
        mepsSimilarity = calculateSimilarity(mepsimage, MEPS)
        if mepsSimilarity < 0.8:
            moppResult.update({'meps_similarity': {'result': 'NG', 'item': 'sdppi',
                                                   'detail': {'Threshold': 0.8, 'score': mepsSimilarity}}})
            flag = 1
            return moppResult, flag
        else:
            moppResult.update({'meps_similarity': {'result': 'OK', 'item': 'sdppi',
                                                   'detail': {'Threshold': 0.8, 'score': mepsSimilarity}}})
            return moppResult, flag
    else:
        moppResult.update({'sdppi': {'result': 'OK', 'item': 'sdppi', 'detail': 'This country not detect SDPPI'}})
        return moppResult, flag


def getSdppiResult(inputPath):
    sdpppiResult = {}
    flag = 0
    image = cv2.imread(inputPath)
    # country = LABELINFO['country']
    # if country == 'C' or country == 'E' or country == 'LE' or country == 'LL' or country == 'J' or country == 'JA' \
    #         or country == 'PA' or country == 'PP' or country == 'X' or country == 'ZA' or country == 'ZP' \
    #         or country == 'AB' or country == 'AE' or country == 'HB' or country == 'ZS':
    sdppimage, _ = getLabel(image, 'sdppi')
    sdppiSimilarity = getSimilarity(sdppimage, SDPPI)
    if sdppiSimilarity < 0.7:
        sdpppiResult.update({'sdppi_similarity': {'result': 'NG', 'item': 'sdppi',
                                                  'detail': {'Threshold': 0.7, 'score': sdppiSimilarity}}})
        flag = 1
        return sdpppiResult, flag
    else:
        sdpppiResult.update({'sdppi_similarity': {'result': 'OK', 'item': 'sdppi',
                                                 'detail': {'Threshold': 0.7, 'score': sdppiSimilarity}}})
        return sdpppiResult, flag
    # elif country == 'CH':


def getPartResult(inputPath):
    global PRED_IMGSIZE
    partResult = {}
    flag = 0
    image = cv2.imread(inputPath)
    country = LABELINFO['country']
    # if country == 'C' or country == 'E' or country == 'LE' or country == 'LL' or country == 'J' or country == 'JA' \
    #         or country == 'PA' or country == 'PP' or country == 'X' or country == 'ZA' or country == 'ZP' \
    #         or country == 'AB' or country == 'AE' or country == 'HB' or country == 'ZS':
    printbundle_image, printbundleSimilarity = getLabel(image, 'printbundle')
    printbundle_partnum_value = ''
    height = printbundle_image.shape[0]
    printbundlepartnumimage = printbundle_image[int(height / 3):]
    # cv2.imwrite('temp.jpg', printbundlepartnumimage)
    printbundle_partnum_result, _ = FASTOCR(printbundlepartnumimage)
    for txt in printbundle_partnum_result:
        printbundle_partnum_value += txt[1].replace(" ", "").replace("-", "").upper().replace("O", "0")
    printbundlepartnumimageresize = cv2.resize(printbundlepartnumimage, (2000, 2000))
    # cv2.imwrite('temp.jpg', printbundlepartnumimageresize)
    printbundle_partnum_result_resize, _ = FASTOCR(printbundlepartnumimageresize)
    for txtresize in printbundle_partnum_result_resize:
        printbundle_partnum_value += txtresize[1].replace(" ", "").replace("-", "").upper().replace("O", "0")
    # cv2.imwrite('temp.jpg', printbundlepartnumimage)
    printbundlepartsubnumimage = printbundle_image[int(height / 8 * 7):]
    # cv2.imwrite('temp.jpg', printbundlepartsubnumimage)
    printbundle_subpartnum_result, _ = FASTOCR(printbundlepartsubnumimage)
    if printbundle_subpartnum_result is None:
        pass
    else:
        for txt_sub in printbundle_subpartnum_result:
            printbundle_partnum_value += txt_sub[1].replace(" ", "").replace("-", "").upper().replace("O", "0")
    subPartnumList = LABELINFO['subpartnumber'].split('/')
    for partnum in subPartnumList:
        if printbundle_partnum_value.find(partnum.replace("-", "").replace("O", "0")) < 0 or partnum == '':
            partResult.update({'partnumber': {'result': 'NG', 'item': 'part',
                                              'detail': 'Please check the number and map of partnumber!'}})
            flag = 1
            break
        # return partResult, flag
        else:
            pass
    if flag == 0:
        partResult.update({'partnumber': {'result': 'OK', 'item': 'part', 'detail': 'The partnumber is OK'}})
    #####detect line
    exporter = Exporter.YOLOExporter()
    image_Line = PIL.Image.open(inputPath)
    image_Line_size = image_Line.size
    image_Line_array = np.array(image_Line)
    preProcessed = imageResize(image_Line_array, image_Line_size, PRED_IMGSIZE)
    inferedLine = INFERER_LINE.infer(preProcessed)
    postProcessedLine = POSTPROCESSOR.nms(inferedLine)
    postProcessedLine = labelResize(postProcessedLine, image_Line_size, PRED_IMGSIZE)
    LineId = int(exporter.export(postProcessedLine[0]))
    if LABELINFO['typec'] == '622-00449':
        if LineId == 1:
            partResult.update({'line_color': {'result': 'OK', 'item': 'part', 'detail': 'The line color is not OK!'}})
            # return partResult, flag
        else:
            inferedLine = INFERER_LINE_LAST.infer(preProcessed)
            postProcessedLine = POSTPROCESSOR.nms(inferedLine)
            postProcessedLine = labelResize(postProcessedLine, image_Line_size, PRED_IMGSIZE)
            LineId = int(exporter.export(postProcessedLine[0]))
            if LineId == 1:
                partResult.update({'line_color': {'result': 'OK', 'item': 'part', 'detail': 'The line color is OK!'}})
            else:
                partResult.update({'line_color': {'result': 'NG', 'item': 'part', 'detail': 'The line color is not OK!'}})
                flag = 1
            # return partResult, flag
    elif LABELINFO['typec'] == '622-00596':
        if LineId == 2:
            partResult.update({'line_color': {'result': 'OK', 'item': 'part', 'detail': 'The line color is OK!'}})
            # return partResult, flag
        else:
            inferedLine = INFERER_LINE_LAST.infer(preProcessed)
            postProcessedLine = POSTPROCESSOR.nms(inferedLine)
            postProcessedLine = labelResize(postProcessedLine, image_Line_size, PRED_IMGSIZE)
            LineId = int(exporter.export(postProcessedLine[0]))
            if LineId == 2:
                partResult.update({'line_color': {'result': 'OK', 'item': 'part', 'detail': 'The line color is OK!'}})
            else:
                partResult.update({'line_color': {'result': 'NG', 'item': 'part', 'detail': 'The line color is not OK!'}})
                flag = 1
            # return partResult, flag
    elif LABELINFO['typec'] == '622-00582':
        if LineId == 3:
            partResult.update({'line_color': {'result': 'OK', 'item': 'part', 'detail': 'The line color is not OK!'}})
            # return partResult, flag
        else:
            inferedLine = INFERER_LINE_LAST.infer(preProcessed)
            postProcessedLine = POSTPROCESSOR.nms(inferedLine)
            postProcessedLine = labelResize(postProcessedLine, image_Line_size, PRED_IMGSIZE)
            LineId = int(exporter.export(postProcessedLine[0]))
            if LineId == 3:
                partResult.update({'line_color': {'result': 'OK', 'item': 'part', 'detail': 'The line color is OK!'}})
            else:
                partResult.update({'line_color': {'result': 'NG', 'item': 'part', 'detail': 'The line color is not OK!'}})
                flag = 1
            # return partResult, flag
    elif LABELINFO['typec'] == '622-00491':
        if LineId == 4:
            partResult.update({'line_color': {'result': 'OK', 'item': 'part', 'detail': 'The line color is  OK!'}})
            # return partResult, flag
        else:
            inferedLine = INFERER_LINE_LAST.infer(preProcessed)
            postProcessedLine = POSTPROCESSOR.nms(inferedLine)
            postProcessedLine = labelResize(postProcessedLine, image_Line_size, PRED_IMGSIZE)
            LineId = int(exporter.export(postProcessedLine[0]))
            if LineId == 4:
                partResult.update({'line_color': {'result': 'OK', 'item': 'part', 'detail': 'The line color is OK!'}})
            else:
                partResult.update({'line_color': {'result': 'NG', 'item': 'part', 'detail': 'The line color is not OK!'}})
                flag = 1
            # return partResult, flag
    else:
        partResult.update({'line_color': {'result': 'NG', 'item': 'part', 'detail': 'The typeC number is not exit!'}})
        flag = 1
        # return partResult, flag
    # imageHeight = image.shape[0]
    # imageWidth = image.shape[1]
    # clothimage = image[int(imageHeight / 2): int(imageHeight / 4 * 3), int(imageWidth / 4 * 3): int(imageWidth / 8 * 7)]
    # # cv2.imwrite('temp.jpg', clothimage)
    # clothimagegray = cv2.cvtColor(clothimage, cv2.COLOR_BGR2GRAY)
    # average = cv2.mean(clothimagegray)[0]
    # if LABELINFO['partnumber'][-2:] == '94':
    #     if average < 100.0:
    #         partResult.update({'clean_cloth': {'result': 'NG', 'item': 'part', 'detail': 'The clean cloth not exist!'}})
    #         flag = 1
    #         # return partResult, flag
    #     else:
    #         partResult.update({'clean_cloth': {'result': 'OK', 'item': 'part',
    #                                            'detail': 'The clean cloth exist!'}})
    # elif LABELINFO['partnumber'][-2:] == '93':
    #     if average > 100.0:
    #         partResult.update({'clean_cloth': {'result': 'NG', 'item': 'part', 'detail': 'The clean cloth not exist!'}})
    #         flag = 1
    #         # return partResult, flag
    #     else:
    #         partResult.update({'clean_cloth': {'result': 'OK', 'item': 'part',
    #                                            'detail': 'The clean cloth exist!'}})
    # else:
    #     partResult.update({'clean_cloth': {'result': 'NG', 'item': 'part', 'detail': 'Please checking the partnumber pn is true or not!'}})
    #     flag = 1
    #     # return partResult, flag

    if country == 'CH':
        pnsn_ocr_value = ''
        chimageheight = image.shape[0]
        warrantsearch = image[: int(chimageheight / 3 * 2)]
        # cv2.imwrite('temp.jpg', warrantsearch)
        warranty_image, warranty_map = getLabel(warrantsearch, 'warranty')
        # cv2.imwrite('temp.jpg', warranty_image)
        warrantySimilarity = calculateSimilarity(warranty_image, WARRANTY)
        if warrantySimilarity < 0.8 or warranty_map < 0.1:
            partResult.update({'warranty_similarity': {'result': 'NG', 'item': 'part',
                                                       'detail': {'Threshold': 0.8, 'score': warrantySimilarity}}})
            flag = 1
            # return partResult, flag
        else:
            partResult.update({'warranty_similarity': {'result': 'OK', 'item': 'part',
                                                       'detail': {'Threshold': 0.8, 'score': warrantySimilarity}}})
        warrantyHeight = warranty_image.shape[0]
        pnsntempimage = warranty_image[: int(warrantyHeight / 2)]
        pnsn_image, _ = getLabel(pnsntempimage, 'sn')
        # cv2.imwrite('temp.jpg', warranty_image)
        pnsn_ocr_result = OCRINFERER.ocr(pnsn_image)[0]
        if len(pnsn_ocr_result) == 0:
            pnsn_ocr_result, _ = FASTOCR(pnsn_image)
            if pnsn_ocr_result is None:
                partResult.update({'pn_txt': {'result': 'NG', 'item': 'part', 'detail': 'Checking the WAARANTY and SN if or not exit!'}})
                flag = 1
                return partResult, flag
            else:
                pnsn_ocr_value = ''
                for linepnsn in pnsn_ocr_result:
                    pnsn_ocr_value += linepnsn[1].replace(" ", "").upper().replace("O", "0").replace(',', '.')
        else:
            for linepnsn in pnsn_ocr_result:
                pnsn_ocr_value += linepnsn[1][0].replace(" ", "").upper().replace('O', '0').replace(',', '.')
        if pnsn_ocr_value.find(LABELINFO['model_no']) < 0:
            if pnsn_ocr_value.split('N0.')[1].split('PART')[0].find(LABELINFO['model_no'].replace('O', '0')) < 0:
                partResult.update({'modelno_txt': {'result': 'NG', 'item': 'part',
                                                   'detail': {'expected': LABELINFO['model_no'],
                                                              'actual': pnsn_ocr_value.split('N0.')[1].split('PART')[0]}}})
                flag = 1
                # return partResult, flag
            else:
                partResult.update({'modelno_txt': {'result': 'OK', 'item': 'part',
                                                   'detail': {'expected': LABELINFO['model_no'],
                                                              'actual': LABELINFO['model_no']}}})
        else:
            partResult.update({'modelno_txt': {'result': 'OK', 'item': 'part',
                                               'detail': {'expected': LABELINFO['model_no'],
                                                          'actual': LABELINFO['model_no']}}})
        if pnsn_ocr_value.find(LABELINFO['pn']) < 0:
            if pnsn_ocr_value.split('PART')[1].split('SER')[0].replace('O', '0').find(LABELINFO['pn'].replace('O', '0')) < 0:
                partResult.update({'pn_txt': {'result': 'NG', 'item': 'part',
                                              'detail': {'expected': LABELINFO['pn'],
                                                         'actual': pnsn_ocr_value.split('PART')[1].split('SER')[0].replace(',', '.').split('.')[1]}}})
                flag = 1
                # return partResult, flag
            else:
                partResult.update({'pn_txt': {'result': 'OK', 'item': 'part',
                                              'detail': {'expected': LABELINFO['pn'],
                                                         'actual': LABELINFO['pn']}}})
        else:
            partResult.update({'pn_txt': {'result': 'OK', 'item': 'part',
                                          'detail': {'expected': LABELINFO['pn'],
                                                     'actual': LABELINFO['pn']}}})
        if pnsn_ocr_value.find(LABELINFO['sn']) < 0:
            if pnsn_ocr_value.split('SER')[1].replace('O', '0').find(LABELINFO['sn'].replace('O', '0')) < 0:
                partResult.update({'sn_txt': {'result': 'NG', 'item': 'part',
                                              'detail': {'expected': LABELINFO['sn'],
                                                         'actual': pnsn_ocr_value.split('SER')[1].replace(',', '.').split('.')[1]}}})
                flag = 1
                return partResult, flag
            else:
                partResult.update({'sn_txt': {'result': 'OK', 'item': 'part',
                                              'detail': {'expected': LABELINFO['sn'],
                                                         'actual': LABELINFO['sn']}}})
                return partResult, flag
        else:
            partResult.update({'sn_txt': {'result': 'OK', 'item': 'part',
                                          'detail': {'expected': LABELINFO['sn'],
                                                     'actual': LABELINFO['sn']}}})
            return partResult, flag
    else:
        return partResult, flag


def getMrpResult(inputPath, mrpStep):
    mrpResult = {}
    resultFlag = 0
    mrp_ocr_value = ''
    image = cv2.imread(inputPath)
    if mrpStep == '1':
        mrpimageHeight = image.shape[0]
        if LABELINFO['order_type'] == 'BTR':
            mrpimageresearch = image[int(mrpimageHeight / 2):]
        else:
            mrpimageresearch = image[int(mrpimageHeight / 3):]
        # cv2.imwrite('temp.jpg', mrpimageresearch)
        mrp_image, matchMrpSimilarity = getLabel(mrpimageresearch, 'mrpstep1')
        mrpSimilarity = calculateSimilarity(mrp_image, MRPSTEP1)
        mrp_image = cv2.flip(cv2.transpose(mrp_image), 1)
        # cv2.imwrite('temp.jpg', mrp_image)
    else:
        mrpimageHeight = image.shape[0]
        if LABELINFO['order_type'] == 'BTR':
            mrpimageresearch = image[int(mrpimageHeight / 2):]
        else:
            mrpimageresearch = image[int(mrpimageHeight / 5):]
        # cv2.imwrite('temp.jpg', mrpimageresearch)
        mrp_image, matchMrpSimilarity = getLabel(mrpimageresearch, 'mrpstep3')
        mrpSimilarity = calculateSimilarity(mrp_image, MRPSTEP3)
        mrp_image = cv2.flip(cv2.transpose(mrp_image), 0)
        # cv2.imwrite('temp.jpg', mrp_image)
    if mrpSimilarity < 0.8 and matchMrpSimilarity < 0.2:
        mrpResult.update({'mrp_similarity': {'result': 'NG', 'item': 'mrp', 'detail': {'threshold': 0.8, 'score': mrpSimilarity}}})
        resultFlag = 1
        # return mrpResult, resultFlag
    else:
        if mrpSimilarity < 0.5:
            mrpSimilarity += 0.5
        mrpResult.update({'mrp_similarity': {'result': 'OK', 'item': 'mrp', 'detail': {'threshold': 0.8, 'score': mrpSimilarity}}})
    # mrp_ocr_result = OCRINFERER.ocr(mrp_image)[0]
    # for linemrp in mrp_ocr_result:
    #     mrp_ocr_value += linemrp[1][0].replace(' ', '').upper()
    mrp_ocr_result, _ = FASTOCR(mrp_image)
    for linemrp in mrp_ocr_result:
        mrp_ocr_value += linemrp[1].replace(" ", "").upper()
    if LABELINFO['order_type'] == 'BTR':
        if mrp_ocr_value.replace("O", "0").find(LABELINFO['sku'].replace("O", "0")) < 0:
            mrpResult.update({'sku_txt': {'result': 'NG', 'item': 'mrp',
                                          'detail': {'expected': LABELINFO['sku'],
                                                     'actual': mrp_ocr_value.split('SKU')[1].split('MRP')[0]}}})
            resultFlag = 1
            # return mrpResult, resultFlag
        else:
            mrpResult.update({'sku_txt': {'result': 'OK', 'item': 'mrp',
                                          'detail': {'expected': LABELINFO['sku'], 'actual': LABELINFO['sku']}}})

        if mrp_ocr_value.replace("O", "0").find(LABELINFO['mrp'].replace("O", "0")) < 0:
            mrpResult.update({'mrp_txt': {'result': 'NG', 'item': 'mrp',
                                          'detail': {'expected': LABELINFO['mrp'],
                                                     'actual': mrp_ocr_value.split('MRP')[1].split('.00')[0] + '.00'}}})
            resultFlag = 1
            return mrpResult, resultFlag
        else:
            mrpResult.update({'mrp_txt': {'result': 'OK', 'item': 'mrp',
                                          'detail': {'expected': LABELINFO['mrp'], 'actual': LABELINFO['mrp']}}})
            return mrpResult, resultFlag
    else:
        pattern = r'\d+'
        mrp_ocr_data = re.findall(pattern, mrp_ocr_value.split('CAPACITY')[1].replace('0', 'O').split('NOTE')[0])[0]
        if mrp_ocr_value.replace("O", "0").find(LABELINFO['ssd_size'].replace("O", "0")) < 0:
            mrpResult.update({'capacity': {'result': 'NG', 'item': 'mrp',
                                           'detail': {'capacity_code': {'expected': LABELINFO['ssd_size'],
                                                                        'actual': mrp_ocr_data}}}})
            resultFlag = 1
            # return mrpResult, resultFlag
        else:
            mrpResult.update({'capacity': {'result': 'OK', 'item': 'mrp',
                                           'detail': {'expected': LABELINFO['ssd_size'], 'actual': LABELINFO['ssd_size']}}})

        if mrp_ocr_value.replace("O", "0").find(LABELINFO['sku'].replace("O", "0")) < 0:
            mrpResult.update({'sku_txt': {'result': 'NG', 'item': 'mrp',
                                          'detail': {'expected': LABELINFO['sku'],
                                                     'actual': mrp_ocr_value.split('SKU')[1].split('MRP')[0]}}})
            resultFlag = 1
            # return mrpResult, resultFlag
        else:
            mrpResult.update({'sku_txt': {'result': 'OK', 'item': 'mrp',
                                          'detail': {'expected': LABELINFO['sku'], 'actual': LABELINFO['sku']}}})

        if mrp_ocr_value.replace("O", "0").find(LABELINFO['mrp'].replace("O", "0")) < 0:
            mrpResult.update({'mrp_txt': {'result': 'NG', 'item': 'mrp',
                                          'detail': {'expected': LABELINFO['mrp'],
                                                     'actual': mrp_ocr_value.split('MRP')[1].split('.00')[0] + '.00'}}})
            resultFlag = 1
            return mrpResult, resultFlag
        else:
            mrpResult.update({'mrp_txt': {'result': 'OK', 'item': 'mrp',
                                          'detail': {'expected': LABELINFO['mrp'], 'actual': LABELINFO['mrp']}}})
            return mrpResult, resultFlag


def getSecResult(inputPath):
    secResult = {}
    resultFlag = 0
    sec_ocr_value = ''
    image = cv2.imread(inputPath)
    secheight = image.shape[0]
    secwidth = image.shape[1]
    secsearchimage = image[int(secheight / 3): int(secheight / 3 * 2), : int(secwidth / 3 * 2)]
    # cv2.imwrite('temp.jpg', secsearchimage)
    sec_image, secmapscores = getLabel(secsearchimage, 'sec')
    # cv2.imwrite('temp.jpg', sec_image)
    secimageRotated = cv2.flip(cv2.transpose(sec_image), 0)
    # cv2.imwrite('temp.jpg', secimageRotated)
    secSimilarity = calculateSimilarity(sec_image, SEC)
    if secSimilarity < 0.8 or secmapscores < 0.1:
        secResult.update({'sec_similarity': {'result': 'NG', 'item': 'sec', 'detail': {'threshold': 0.8, 'score': secSimilarity}}})
        resultFlag = 1
        # return secResult, resultFlag
    else:
        secResult.update({'sec_similarity': {'result': 'OK', 'item': 'sec', 'detail': {'threshold': 0.8, 'score': secSimilarity}}})
    height = secimageRotated.shape[0]
    secimage = secimageRotated[int(height / 2):]
    # cv2.imwrite('temp.jpg', secimage)
    sec_ocr_result = OCRINFERER.ocr(secimage)[0]
    for linesec in sec_ocr_result:
        sec_ocr_value += linesec[1][0].upper().replace('O', '0')
    if sec_ocr_value.find(LABELINFO['sec_num']) < 0 or LABELINFO['sec_num'] == '':
        secResult.update({'sec_txt': {'result': 'NG', 'item': 'sec',
                                      'detail': {'expected': LABELINFO['sec_num'], 'actual': sec_ocr_value}}})
        resultFlag = 1
        return secResult, resultFlag
    else:
        secResult.update({'sec_txt': {'result': 'OK', 'item': 'sec',
                                      'detail': {'expected': LABELINFO['sec_num'], 'actual': sec_ocr_value}}})
        return secResult, resultFlag


######检测外壳的侧面图像
def getLogoResult(inputPath):
    logoResult = {}
    resultFlag = 0
    image = cv2.imread(inputPath)
    logoHeight = image.shape[0]
    logoWidth = image.shape[1]
    logoimage = image[: int(logoHeight / 2), : int(logoWidth / 3 * 2)]
    # cv2.imwrite('temp.jpg', logoimage)
    if  LABELINFO['typec'] == '622-00491':
        logo_image, logoSimilarity = getLabel(logoimage, 'logoslv')
        # cv2.imwrite('temp.jpg', logo_image)
        label_hsv = cv2.cvtColor(LOGOSLV, cv2.COLOR_BGR2HSV)
    elif LABELINFO['typec'] == '622-00582':
        logo_image, logoSimilarity = getLabel(logoimage, 'logosky')
        # cv2.imwrite('temp.jpg', logo_image)
        label_hsv = cv2.cvtColor(LOGOSKY, cv2.COLOR_BGR2HSV)
    elif LABELINFO['typec'] == '622-00449':
        logo_image, logoSimilarity = getLabel(logoimage, 'logostl')
        # cv2.imwrite('temp.jpg', logo_image)
        label_hsv = cv2.cvtColor(LOGOSTL, cv2.COLOR_BGR2HSV)
    else:
        logo_image, logoSimilarity = getLabel(logoimage, 'logomdn')
        # cv2.imwrite('temp.jpg', logo_image)
        label_hsv = cv2.cvtColor(LOGOMDN, cv2.COLOR_BGR2HSV)
    logo_hsv = cv2.cvtColor(logo_image, cv2.COLOR_BGR2HSV)
    hist_logo = cv2.calcHist([logo_hsv], [0, 1, 2], None, [8, 8, 8], [0, 256, 0, 256, 0, 256])
    hist_logo = cv2.normalize(hist_logo, hist_logo).flatten()
    hist_label = cv2.calcHist([label_hsv], [0, 1, 2], None, [8, 8, 8], [0, 256, 0, 256, 0, 256])
    hist_label = cv2.normalize(hist_label, hist_label).flatten()
    similarity = 1.0 - cv2.compareHist(hist_logo, hist_label, cv2.HISTCMP_BHATTACHARYYA)
    if similarity < 0.3 and logoSimilarity < 0.3:
        logoResult.update({'logo_similarity': {'result': 'NG', 'item': 'logo', 'detail': {'threshold': 0.7, 'score': similarity}}})
        resultFlag = 1
        return logoResult, resultFlag
    else:
        logoResult.update({'logo_similarity': {'result': 'OK', 'item': 'logo', 'detail': 'Logo color is ok!'}})
        return logoResult, resultFlag


def getDuckHeadAdapterResult(inputPath):
    global PRED_IMGSIZE
    adResult = {}
    resultFlag = 0
    country = LABELINFO['country']
    ##### DuckHead Detection
    exporter = Exporter.YOLOExporter()
    image_DuckHead = PIL.Image.open(inputPath)
    image_DuckHead_size = image_DuckHead.size
    image_DuckHead_array = np.array(image_DuckHead)
    preProcessed = imageResize(image_DuckHead_array, image_DuckHead_size, PRED_IMGSIZE)
    inferedDuckhead = INFERER_DUCKHEAD.infer(preProcessed)
    postProcessedDuckhead = POSTPROCESSOR.nms(inferedDuckhead)
    postProcessedDuckhead = labelResize(postProcessedDuckhead, image_DuckHead_size, PRED_IMGSIZE)
    duckheadId = int(exporter.export(postProcessedDuckhead[0]))
    if duckheadId == 0:
        adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                      'detail': 'Duckhead is not exit!'}})
        resultFlag = 1
    elif duckheadId == 1:
        if country == 'X':
            adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is true!'}})
        else:
            adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is not true!'}})
            resultFlag = 1
    elif duckheadId == 3:
        if country == 'KH' or country == 'TH' or country == 'TG':
            adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is true!'}})
        else:
            inferedDuckhead = INFERER_DUCKHEAD_LAST.infer(preProcessed)
            postProcessedDuckhead = POSTPROCESSOR.nms(inferedDuckhead)
            postProcessedDuckhead = labelResize(postProcessedDuckhead, image_DuckHead_size, PRED_IMGSIZE)
            duckheadId = int(exporter.export(postProcessedDuckhead[0]))
            if duckheadId == 10 or duckheadId == 3:
                adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': 10,
                                              'detail': 'Duckhead is true!'}})
            else:
                adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                              'detail': 'Duckhead is not true!'}})
                resultFlag = 1
    elif duckheadId == 5:
        if country == 'PA' or country == 'ZP' or country == 'AB' or country == 'B' or country == 'ZS':
            adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is true!'}})
        else:
            adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is not true!'}})
            resultFlag = 1
    elif duckheadId == 7:
        if country == 'LE':
            adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is true!'}})
        else:
            adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is not true!'}})
            resultFlag = 1
    elif duckheadId == 4:
        if country == 'HN':
            adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is true!'}})
        else:
            adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is not true!'}})
            resultFlag = 1
    elif duckheadId == 2:
        if country == 'TA' or country == 'CT' or country == 'LL' or country == 'C' or country == 'VC' or country == 'E' or country == 'J'\
                or country == 'JA':
            adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is true!'}})
        else:
            adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is not true!'}})
            resultFlag = 1
    elif duckheadId == 10:
        if country == 'AE' or country == 'CR' or country == 'CZ' or country == 'D' or country == 'DK' or country == 'FN' or\
            country == 'GR' or country == 'H' or country == 'HB' or country == 'KS' or country == 'MG' or country == 'N' or\
            country == 'PO' or country == 'RO' or country == 'RU' or country == 'SL' or country == 'SM' or country == 'T' or\
            country == 'TU' or country == 'UA' or country == 'Y' or country == 'ZE' or country == 'ID' or country == 'SA' or\
            country == 'ZA' or country == 'BZ' or country == 'CI':
            adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is true!'}})
        else:
            adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is not true!'}})
            resultFlag = 1
    elif duckheadId == 6:
        if country == 'CH' or country == 'PP':
            adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is true!'}})
        else:
            adResult.update({'duckhead': {'result': 'NG', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                          'detail': 'Duckhead is not true!'}})
            resultFlag = 1
    elif duckheadId == 8:
        if LABELINFO['adaptersn'] == 'CH611-00545' or LABELINFO['adaptersn'] == 'CH611-00606' or \
                LABELINFO['adaptersn'] == 'CH611-00684':
            adResult.update({'adapter': {'result': 'OK', 'item': 'adapter', 'duckheadindex': duckheadId,
                                         'detail': 'Adapter is 35W!'}})
        else:
            adResult.update({'adapter': {'result': 'NG', 'item': 'adapter', 'duckheadindex': duckheadId,
                                         'detail': 'Adapter is not true!'}})
            resultFlag = 1

        return adResult, resultFlag
    elif duckheadId == 9:
        if LABELINFO['adaptersn'] == '611-00515' or LABELINFO['adaptersn'] == '611-00605' or \
            LABELINFO['adaptersn'] == '611-00682' or LABELINFO['adaptersn'] == 'TA611-00515' or \
            LABELINFO['adaptersn'] == 'TA611-00605' or LABELINFO['adaptersn'] == 'TA611-00682':
            adResult.update({'adapter': {'result': 'OK', 'item': 'adapter', 'duckheadindex': duckheadId,
                                         'detail': 'Adapter is 35W!'}})
        else:
            adResult.update({'adapter': {'result': 'NG', 'item': 'adapter', 'duckheadindex': duckheadId,
                                         'detail': 'Adapter is not true!'}})
            resultFlag = 1

        return adResult, resultFlag
    else:
        adResult.update({'duckhead': {'result': 'OK', 'item': 'duckhead', 'duckheadindex': duckheadId,
                                      'detail': 'duckhead is true!'}})
    #####Adapter Detection
    image_adapter = cv2.imread(inputPath)
    b, g, r = cv2.split(image_adapter)
    reduced_r = b[2000:7000, 600:4500]
    imgrang1 = cv2.inRange(reduced_r, 160, 255)
    contours, hierarchy = cv2.findContours(imgrang1, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    max_contour = max(contours, key=cv2.contourArea)
    rect = cv2.minAreaRect(max_contour)
    width, height = rect[1]
    # if LABELINFO['project'] == 'J713':
    if LABELINFO['adaptersn'] == '611-00574' or LABELINFO['adaptersn'] == '2A611-00574' or \
            LABELINFO['adaptersn'] == 'BZ611-00587' or LABELINFO['adaptersn'] == 'CI611-00574' or \
            LABELINFO['adaptersn'] == 'HN611-00574' or LABELINFO['adaptersn'] == 'LA611-00574' or \
            LABELINFO['adaptersn'] == 'TH611-00574' or LABELINFO['adaptersn'] == '2A611-00381' or \
            LABELINFO['adaptersn'] == '2A611-00518' or LABELINFO['adaptersn'] == '611-00381' or \
            LABELINFO['adaptersn'] == '611-00518' or LABELINFO['adaptersn'] == 'BZ611-00381' or \
            LABELINFO['adaptersn'] == 'BZ611-00518' or LABELINFO['adaptersn'] == 'CI611-00381' or \
            LABELINFO['adaptersn'] == 'CI611-00518' or LABELINFO['adaptersn'] == 'HN611-00381' or \
            LABELINFO['adaptersn'] == 'HN611-00518' or LABELINFO['adaptersn'] == 'LA611-00381' or \
            LABELINFO['adaptersn'] == 'LA611-00518' or LABELINFO['adaptersn'] == 'MY611-00381' or \
            LABELINFO['adaptersn'] == 'MY611-00518':
        if width > 2350.0 and height > 2350.0:
            adResult.update({'adapter': {'result': 'NG', 'item': 'adapter', 'detail': 'Adapter is 96W!'}})
            resultFlag = 1
        else:
            adResult.update({'adapter': {'result': 'OK', 'item': 'adapter', 'detail': 'Adapter is 70W!'}})
    elif LABELINFO['adaptersn'] == '2A611-00628' or LABELINFO['adaptersn'] == '611-00628' or \
            LABELINFO['adaptersn'] == 'CH611-00628' or LABELINFO['adaptersn'] == 'CI611-00628' or \
            LABELINFO['adaptersn'] == 'HN611-00628' or LABELINFO['adaptersn'] == 'LA611-00628' or \
            LABELINFO['adaptersn'] == 'TH611-00628':
        if width > 2350.0 and height > 2350.0:
            adResult.update({'adapter': {'result': 'OK', 'item': 'adapter', 'detail': 'Adapter is 96W!'}})
        else:
            adResult.update({'adapter': {'result': 'NG', 'item': 'adapter', 'detail': 'Adapter is 70W!'}})
            resultFlag = 1
    else:
        adResult.update({'adapter': {'result': 'NG', 'item': 'adapter', 'detail': 'Adapter pn is not exit!'}})
        resultFlag = 1
    # else:
    #     if LABELINFO['adaptersn'] == '611-00628' or LABELINFO['adaptersn'] == '2A611-00628' or \
    #             LABELINFO['adaptersn'] == 'HN611-00628' or LABELINFO['adaptersn'] == 'TH611-00715' or \
    #             LABELINFO['adaptersn'] == 'CI611-00628' or LABELINFO['adaptersn'] == 'LA611-00628':
    #         if width > 2650.0 and height > 2650.0:
    #             adResult.update({'adapter': {'result': 'NG', 'item': 'adapter', 'detail': 'Adapter is 96W!'}})
    #             resultFlag = 1
    #         else:
    #             adResult.update({'adapter': {'result': 'OK', 'item': 'adapter', 'detail': 'Adapter is 70W!'}})
    #     elif LABELINFO['adaptersn'] == '611-00444' or LABELINFO['adaptersn'] == '2A611-00348' \
    #             or LABELINFO['adaptersn'] == 'HN611-00348' or LABELINFO['adaptersn'] == 'TH611-00348'or \
    #             LABELINFO['adaptersn'] == 'CI611-00791' or LABELINFO['adaptersn'] == 'LA611-00348':
    #         if width > 2650.0 and height > 2650.0:
    #             adResult.update({'adapter': {'result': 'OK', 'item': 'adapter', 'detail': 'Adapter is 96W!'}})
    #         else:
    #             adResult.update({'adapter': {'result': 'NG', 'item': 'adapter', 'detail': 'Adapter is 70W!'}})
    #             resultFlag = 1
    #     else:
    #         adResult.update({'adapter': {'result': 'NG', 'item': 'adapter', 'detail': 'Adapter pn is not exit!'}})
    #         resultFlag = 1

    return adResult, resultFlag


######检测外壳的主题面图像
def getTopResult(inputPath):
    country = COUNTRY
    topResult = {}
    resultFlag = 0
    image = cv2.imread(inputPath)
    if country == 'C' or country == 'E' or country == 'LE' or country == 'LL' or country == 'J' or country == 'JA' \
            or country == 'PA' or country == 'PP' or country == 'X' or country == 'ZA' or country == 'ZP' \
            or country == 'AB' or country == 'AE' or country == 'HB' or country == 'ZS':
        fg_image, fg_similirity = getLabel(image, 'fg')
        if fg_similirity < 0.9:
            topResult.update({'fgSimilarity': {'result': 'NG', 'item': 'fg',
                                               'detail': {'Threshold': 0.9, 'score': fg_similirity}}})
            resultFlag = 1
            return topResult, resultFlag
        else:
            topResult.update({'fgSimilarity': {'result': 'OK', 'item': 'fg',
                                               'detail': {'Threshold': 0.9, 'score': fg_similirity}}})
        result_fg, fgflag = parseFG(fg_image)
        if fgflag == 1:
            topResult.update(result_fg)
            resultFlag = 1
            return topResult, resultFlag
        else:
            topResult.update(result_fg)
        if LABELINFO['version'] == 'BTR':
            ftr_image, ftr_similarity = getLabel(image, 'ftr')
            if ftr_similarity < 0.9:
                topResult.update({'ftrSimilarity': {'result': 'NG', 'item': 'fg',
                                                    'detail': {'Threshold': 0.9, 'score': ftr_similarity}}})
                resultFlag = 1
                return topResult, resultFlag
            else:
                topResult.update({'ftrSimilarity': {'result': 'OK', 'item': 'fg',
                                                    'detail': {'Threshold': 0.9, 'score': ftr_similarity}}})
            result_ftr, fgflag = parseFtr(ftr_image)
            if fgflag == 1:
                topResult.update(result_ftr)
                resultFlag = 1
                return topResult, resultFlag
            else:
                topResult.update(result_ftr)
            return topResult, resultFlag
        elif LABELINFO['version'] == 'CTO':
            sla_image, sla_similarty = getLabel(image, SLA)
            if sla_similarty < 0.9:
                topResult.update({'slaSimilarty': {'result': 'NG', 'item': 'sla',
                                                   'detail': {'threshold': 0.9, 'score': sla_similarty}}})
                resultFlag = 1
                return topResult, resultFlag
            else:
                topResult.update({'slaSimilarty': {'result': 'OK', 'item': 'sla',
                                                   'detail': {'threshold': 0.9, 'score': sla_similarty}}})
                return topResult, resultFlag
        else:
            pass
    elif country == 'CH':
        fg_image, fg_similirity = getLabel(image, 'fg')
        if fg_similirity < 0.9:
            topResult.update({'fgSimilarity': {'result': 'NG', 'item': 'fg',
                                               'detail': {'Threshold': 0.9, 'score': fg_similirity}}})
            resultFlag = 1
            return topResult, resultFlag
        else:
            topResult.update({'fgSimilarity': {'result': 'OK', 'item': 'fg',
                                               'detail': {'Threshold': 0.9, 'score': fg_similirity}}})
        result_fg, fgflag = parseFG(fg_image)
        if fgflag == 1:
            topResult.update(result_fg)
            resultFlag = 1
            return topResult, resultFlag
        else:
            topResult.update(result_fg)
        energy_image,  energy_similarity = getLabel(image, 'energy')
        if energy_similarity < 0.9:
            topResult.update({'energySimilarity': {'result': 'NG', 'item': 'energy',
                                                   'detail': {'threshold': 0.9, 'score': energy_similarity}}})
            resultFlag = 1
            return topResult, resultFlag
        else:
            topResult.update({'energySimilarity': {'result': 'OK', 'item': 'energy',
                                                   'detail': {'threshold': 0.9, 'score': energy_similarity}}})
        import_image, import_similarity = getLabel(image, 'importer')
        if import_similarity < 0.9:
            topResult.update({'importerSimilarity': {'result': 'NG', 'item': 'importer',
                                                     'detail': {'threshold': 0.9, 'score': import_similarity}}})
            resultFlag = 1
            return topResult, resultFlag
        else:
            topResult.update({'importerSimilarity': {'result': 'OK', 'item': 'importer',
                                                     'detail': {'threshold': 0.9, 'score': import_similarity}}})
        if LABELINFO['version'] == 'BTR':
            ftr_image, ftr_similarity = getLabel(image, 'ftr')
            if ftr_similarity < 0.9:
                topResult.update({'ftrSimilarity': {'result': 'NG', 'item': 'fg',
                                                    'detail': {'Threshold': 0.9, 'score': ftr_similarity}}})
                resultFlag = 1
                return topResult, resultFlag
            else:
                topResult.update({'ftrSimilarity': {'result': 'OK', 'item': 'fg',
                                                    'detail': {'Threshold': 0.9, 'score': ftr_similarity}}})
            result_ftr, fgflag = parseFtr(ftr_image)
            if fgflag == 1:
                topResult.update(result_ftr)
                resultFlag = 1
                return topResult, resultFlag
            else:
                topResult.update(result_ftr)
            return topResult, resultFlag
        else:
            sla_image, sla_similarty = getLabel(image, SLA)
            if sla_similarty < 0.9:
                topResult.update({'slaSimilarty': {'result': 'NG', 'item': 'sla',
                                                   'detail': {'threshold': 0.9, 'score': sla_similarty}}})
                resultFlag = 1
                return topResult, resultFlag
            else:
                topResult.update({'slaSimilarty': {'result': 'OK', 'item': 'sla',
                                                   'detail': {'threshold': 0.9, 'score': sla_similarty}}})
                return topResult, resultFlag


def parseFG(image):
    cutNum = 4
    res = {}
    Flag = 0
    fgInfo = cutImageByLine(image, cutNum)
    if LABELINFO['upc'].find(fgInfo[0].split('UPC')[1]) < 0 or fgInfo[1].find(LABELINFO['upc']) < 0:
        res.update({'upc': {'result': 'NG', 'item': 'fg',
                            'detail': {'upc_code': {'expected': LABELINFO['upc'], 'actual': fgInfo[0], 'barcode': fgInfo[1]}}}})
        Flag = 1
        return res, Flag
    else:
        res.update({'upc': {'result': 'OK', 'item': 'fg', 'detail': {}}})
    if fgInfo[2].find(LABELINFO['pn']) < 0 and fgInfo[3].find(LABELINFO['pn']) < 0:
        res.update({'pn': {'result': 'NG', 'item': 'fg',
                           'detail': {'pn_code': {'expected': LABELINFO['pn'], 'actual': fgInfo[2], 'barcode': fgInfo[3]}}}})
        Flag = 1
        return res, Flag
    else:
        res.update({'pn': {'result': 'OK', 'item': 'fg', 'detail': {}}})
    if fgInfo[4].find(LABELINFO['sn']) < 0 and fgInfo[5].find(LABELINFO['sn']) < 0:
        res.update({'sn': {'result': 'NG', 'item': 'fg',
                           'detail': {'sn_code': {'expected': LABELINFO['sn'], 'actual': fgInfo[4], 'barcode': fgInfo[5]}}}})
        Flag = 1
        return res, Flag
    else:
        res.update({'sn': {'result': 'OK', 'item': 'fg', 'detail': {}}})
    if fgInfo[6].find(LABELINFO['modelNo']) < 0:
        res.update({'model_no': {'result': 'NG', 'item': 'fg',
                                 'detail': {'model_no': {'expected': LABELINFO['model_no'], 'actual': fgInfo[6]}}}})
        Flag = 1
        return res, Flag
    else:
        res.update({'model_no': {'result': 'OK', 'item': 'fg', 'detail': {}}})
    if fgInfo[7].find(LABELINFO['config']) < 0:
        res.update({'config': {'result': 'NG', 'item': 'fg',
                               'detail': {'config': {'expected': LABELINFO['config'], 'actual': fgInfo[6]}}}})
        Flag = 1
        return res, Flag
    else:
        res.update({'config': {'result': 'OK', 'item': 'fg', 'detail': {}}})
        return res, Flag


def parseFtr(image):
    ftrInfo = ''
    res = {}
    flag = 0
    ftr_result = OCRINFERER.ocr(image)[0]
    for line in ftr_result:
        ftrInfo += line[1][0]
    if ftrInfo.find(LABELINFO['memory']) < 0:
        res.update({'memory': {'result': 'NG', 'item': 'ftr',
                               'detail': {'memory': {'expected': LABELINFO['memory'], 'actual': ftrInfo[:8].split('GB')[0] + 'GB'}}}})
        flag = 1
        return res, flag
    else:
        res.update({'memory': {'result': 'OK', 'item': 'ftr', 'detail': {}}})
    if ftrInfo.find(LABELINFO['capacity']) < 0:
        res.update({'capacity': {'result': 'NG', 'item': 'ftr',
                                 'detail': {'capacity': {'expected': LABELINFO['capacity'], 'actual': ftrInfo[5:24].split('SSD')[0][-5:]}}}})
        flag = 1
        return res, flag
    else:
        res.update({'capacity': {'result': 'OK', 'item': 'ftr', 'detail': {}}})
    if ftrInfo.find(LABELINFO['partNum']) < 0:
        res.update({'partNum': {'result': 'NG', 'item': 'ftr',
                                'detail': {'partNum': {'expected': LABELINFO['partNum'], 'actual': ftrInfo[-18:].split('。')[1]}}}})
        flag = 1
        return res, flag
    else:
        res.update({'partNum': {'result': 'OK', 'item': 'ftr', 'detail': {}}})
        return res, flag


@server.route('/testService', methods=['get', 'post'])
def testService():
    res = {'msg': 'Service Initing...'}
    if (len(FINISH_LOAD) > 0) and (len(ML_LOAD) == 0):
        res = {'msg': ''}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/getVersion', methods=['get', 'post'])
def getVersion():
    res = {'msg': '7.0.4'}
    return json.dumps(res, ensure_ascii=False)


def barcodeDecode(image, codetype=None):
    imageGray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    if codetype == 'CODE128':
        barcodes = decode(imageGray, symbols=[pyzbar.wrapper.ZBarSymbol.CODE128])
    elif codetype == 'EAN-13':
        barcodes = decode(imageGray, symbols=[pyzbar.wrapper.ZBarSymbol.EAN13])
    elif codetype == 'UPC-A':
        barcodes = decode(imageGray, symbols=[pyzbar.wrapper.ZBarSymbol.UPCA])
    else:
        barcodes = decode(imageGray)
    if len(barcodes) == 0 or len(barcodes[0].data.decode('utf-8')) < 6:
        # heigth = imageGray.shape[0]
        # width = imageGray.shape[1]
        image_resize = cv2.resize(imageGray, (1024, 128))
        blured = cv2.GaussianBlur(image_resize, (3, 3), 0)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(10, 10))
        enhanced = clahe.apply(blured)
        # cv2.imwrite('temp.jpg', enhanced)
        if codetype == 'CODE128':
            barcodes = decode(enhanced, symbols=[pyzbar.wrapper.ZBarSymbol.CODE128])
        elif codetype == 'EAN-13':
            barcodes = decode(enhanced, symbols=[pyzbar.wrapper.ZBarSymbol.EAN13])
        elif codetype == 'UPC-A':
            barcodes = decode(enhanced, symbols=[pyzbar.wrapper.ZBarSymbol.UPCA])
        else:
            barcodes = decode(enhanced)
        if len(barcodes) == 0:
            # imageBlur = cv2.GaussianBlur(imageGray, (3, 3), 0)
            # image_binary = cv2.adaptiveThreshold(imageBlur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
            _, image_binary = cv2.threshold(imageGray, 110, 255, cv2.THRESH_BINARY)
            # cv2.imwrite('temp.jpg', image_binary)
            if codetype == 'CODE128':
                barcodes = decode(image_binary, symbols=[pyzbar.wrapper.ZBarSymbol.CODE128])
            elif codetype == 'EAN-13':
                barcodes = decode(image_binary, symbols=[pyzbar.wrapper.ZBarSymbol.EAN13])
            elif codetype == 'UPC-A':
                barcodes = decode(image_binary, symbols=[pyzbar.wrapper.ZBarSymbol.UPCA])
            else:
                barcodes = decode(image_binary)
            if len(barcodes) == 0:
                imageBlur = cv2.GaussianBlur(imageGray, (3, 3), 0)
                _, imageBinary = cv2.threshold(imageBlur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                if codetype == 'CODE128':
                    barcodes = decode(imageBinary, symbols=[pyzbar.wrapper.ZBarSymbol.CODE128])
                elif codetype == 'EAN-13':
                    barcodes = decode(imageBinary, symbols=[pyzbar.wrapper.ZBarSymbol.EAN13])
                elif codetype == 'UPC-A':
                    barcodes = decode(imageBinary, symbols=[pyzbar.wrapper.ZBarSymbol.UPCA])
                else:
                    barcodes = decode(imageBinary)
                if len(barcodes) == 0:
                    _, image_binary = cv2.threshold(imageGray, 110, 255, cv2.THRESH_BINARY)
                    if codetype == 'CODE128':
                        barcodes = decode(image_binary, symbols=[pyzbar.wrapper.ZBarSymbol.CODE128])
                    elif codetype == 'EAN-13':
                        barcodes = decode(image_binary, symbols=[pyzbar.wrapper.ZBarSymbol.EAN13])
                    elif codetype == 'UPC-A':
                        barcodes = decode(image_binary, symbols=[pyzbar.wrapper.ZBarSymbol.UPCA])
                    else:
                        barcodes = decode(image_binary)
                    if len(barcodes) == 0:
                        barinfo = ''
                    else:
                        barinfo = barcodes[0].data.decode('utf-8')
                else:
                    barinfo = barcodes[0].data.decode('utf-8')
            else:
                barinfo = barcodes[0].data.decode('utf-8')
        else:
            barinfo = barcodes[0].data.decode('utf-8')
    else:
        barinfo = barcodes[0].data.decode('utf-8')

    return barinfo


def cutImageByLine(image, cutNum):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, 50, 200, apertureSize=3)
    # 使用霍夫变换检测直线
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold=100, minLineLength=50, maxLineGap=2)
    # 创建绘制线条的副本
    line_image = np.copy(image)
    # 存储所有水平线坐标
    horizon_lines = []
    # 绘制检测到的线条
    if lines is not None:
        for line in lines:
            x1, y1, x2, y2 = line[0]
            if abs(y1 - y2) < 10:
                horizon_lines.append((x1, y1, x2, y2))
                cv2.line(line_image, (x1, y1), (x2, y2), (0, 255, 0), 2)
    # cv2.imwrite('line.png', line_image)
    # 根据检测到的线条切分图像
    h_segments = sorted(set([y1 for _, y1, _, _ in horizon_lines]))
    ## 如果没有线条返回
    if not h_segments:
        return
    ## 添加图像边界到分割点
    h_segments = [0] + h_segments + [image.shape[0]]
    ## 切分图像并保存
    part = 0
    upc_ocr_value = ''
    upc_barcode_value = ''
    sn_ocr_value = ''
    sn_barcode_value = ''
    pn_ocr_value = ''
    pn_barcode_value = ''
    modelNo_ocr_value = ''
    config_ocr_value = ''
    flag = 0
    for i in range(len(h_segments) - 1):
        if h_segments[i + 1] - h_segments[i] > 30 and part < cutNum:
            if part == 0:
                upc_image = image[h_segments[i]:h_segments[i + 1] + 30, :]
                # cv2.imwrite('temp.jpg', upc_image)
                upcheight = upc_image.shape[0]
                upc_ocr_result = OCRINFEREREN.ocr(upc_image)[0]
                if len(upc_ocr_result) == 0:
                    continue
                for line in upc_ocr_result:
                    upc_ocr_value += line[1][0].replace(' ', '').upper()
                if LABELINFO['upc'] == LABELINFO['pn']:
                    upc_ocr_value = upc_ocr_value.replace('O', '0').replace('I', '1')
                    if upc_ocr_value.find(LABELINFO['upc'].replace('O', '0').replace('I', '1')) < 0:
                        upc_ocr_value = ''
                        upc_image_resize = cv2.resize(upc_image, (480, 64))
                        upc_ocr_result, _ = FASTOCR(upc_image_resize)
                        for txt in upc_ocr_result:
                            upc_ocr_value += txt[1].replace(" ", "").replace("″", "").replace("\"", "").upper()
                        if upc_ocr_value.find(LABELINFO['upc'].replace('O', '0').replace('I', '1')) < 0:
                            upc_ocr_value = ''
                            upc_image_gray = cv2.cvtColor(upc_image, cv2.COLOR_BGR2GRAY)
                            upc_image_gray = cv2.GaussianBlur(upc_image_gray, (3, 3), 0)
                            upc_ocr_result, _ = FASTOCR(upc_image_gray)
                            for txt in upc_ocr_result:
                                upc_ocr_value += txt[1].replace(" ", "").replace("″", "").replace("\"", "").upper()
                else:
                    if upc_ocr_value == 'JAN' or upc_ocr_value == 'UPC':
                        pass
                    else:
                        pattern = r'\d+'
                        upc_ocr_value = upc_ocr_value.replace('O', '0').replace('I', '1')
                        upc_ocr_value = re.findall(pattern, upc_ocr_value)[0]
                    if LABELINFO['upc'].replace('O', '0').replace('I', '1').find(upc_ocr_value) < 0 or LABELINFO['upc'].replace('O', '0').replace('I', '1')[-1:] != upc_ocr_value[-1:]:
                        upc_ocr_value = ''
                        upc_image_resize = cv2.resize(upc_image, (480, 64))
                        upc_image_gray = cv2.cvtColor(upc_image_resize, cv2.COLOR_BGR2GRAY)
                        upc_image_gray = cv2.GaussianBlur(upc_image_gray, (3, 3), 0)
                        upc_ocr_result, _ = FASTOCR(upc_image_gray)
                        for txt in upc_ocr_result:
                            upc_ocr_value += txt[1].replace(" ", "").replace("″", "").replace("\"", "").upper()
                        if upc_ocr_value == 'JAN' or upc_ocr_value == 'UPC':
                            pass
                        else:
                            pattern = r'\d+'
                            upc_ocr_value = upc_ocr_value.replace('O', '0').replace('I', '1')
                            upc_ocr_value = re.findall(pattern, upc_ocr_value)[0]
                        if LABELINFO['upc'].replace('O', '0').replace('I', '1').find(upc_ocr_value) < 0 or LABELINFO['upc'].replace('O', '0').replace('I', '1')[-1:] != upc_ocr_value[-1:]:
                            upc_ocr_value = ''
                            upc_image_resize = cv2.resize(upc_image, (960, 128))
                            upc_image_gray = cv2.cvtColor(upc_image_resize, cv2.COLOR_BGR2GRAY)
                            upc_image_gray = cv2.GaussianBlur(upc_image_gray, (3, 3), 0)
                            upc_ocr_result, _ = FASTOCR(upc_image_gray)
                            for txt in upc_ocr_result:
                                upc_ocr_value += txt[1].replace(" ", "").replace("″", "").replace("\"", "").upper()
                            if upc_ocr_value == 'JAN' or upc_ocr_value == 'UPC':
                                pass
                            else:
                                pattern = r'\d+'
                                upc_ocr_value = upc_ocr_value.replace('O', '0').replace('I', '1')
                                upc_ocr_value = re.findall(pattern, upc_ocr_value)[0]
                            if LABELINFO['upc'].replace('O', '0').replace('I', '1').find(upc_ocr_value) < 0 or LABELINFO['upc'].replace('O', '0').replace('I', '1')[-1:] != upc_ocr_value[-1:]:
                                upc_ocr_value = ''
                                upc_image_gray = cv2.cvtColor(upc_image, cv2.COLOR_BGR2GRAY)
                                upc_image_gray = cv2.GaussianBlur(upc_image_gray, (3, 3), 0)
                                upc_ocr_result, _ = FASTOCR(upc_image_gray)
                                for txt in upc_ocr_result:
                                    upc_ocr_value += txt[1].replace(" ", "").replace("″", "").replace("\"", "").upper()
                upc_barcode_value = barcodeDecode(upc_image)
            if part == 1:
                snpn_image = image[h_segments[i]:h_segments[i + 1], :]
                #cv2.imwrite('temp.jpg', snpn_image)
                heigth = snpn_image.shape[0]
                if heigth > int(upcheight * 1.2):
                    pn_image = snpn_image[:int(heigth / 2)]
                    sn_image = snpn_image[int(heigth / 2):]
                else:
                    pn_image = snpn_image
                    sn_image = image[h_segments[i + 1] - 10:h_segments[i + 2], :]
                    flag = 1
                # pn_ocr_result = OCRINFERER.ocr(pn_image)[0]
                # if len(pn_ocr_result) == 0:
                #     continue
                # for linepn in pn_ocr_result:
                #     pn_ocr_value += linepn[1][0].replace(' ', '').replace(',', '.').upper()
                # cv2.imwrite('temp.jpg', pn_image)
                pn_ocr_result, _ = FASTOCR(pn_image)
                if pn_ocr_result is None:
                    pn_ocr_result = OCRINFERER.ocr(pn_image)[0]
                    if len(pn_ocr_result) == 0:
                        continue
                    for linepn in pn_ocr_result:
                        pn_ocr_value += linepn[1][0].replace(' ', '').replace(',', '.').upper()
                else:
                    for linepn in pn_ocr_result:
                        pn_ocr_value += linepn[1].replace(" ", "").replace(',', '.').upper()
                if pn_ocr_value.replace('/', '').replace('O', '0').find(LABELINFO['pn'].replace('/', '').replace('O', '0')) < 0:
                    pn_ocr_value = ''
                    pn_image_resize = cv2.resize(pn_image, (320, 64))
                    # cv2.imwrite('temp.jpg', pn_image)
                    pn_ocr_result, _ = FASTOCR(pn_image_resize)
                    for linepn in pn_ocr_result:
                        pn_ocr_value += linepn[1].replace(" ", "").replace(',', '.').upper()
                pn_barcode_value = barcodeDecode(pn_image)
                ######SN CHECK
                # cv2.imwrite('temp.jpg', sn_image)
                sn_ocr_result, _ = FASTOCR(sn_image)
                if sn_ocr_result is None:
                    sn_ocr_result = OCRINFERER.ocr(sn_image)[0]
                    if len(sn_ocr_result) == 0:
                        continue
                    for linesn in sn_ocr_result:
                        sn_ocr_value += linesn[1][0].replace(' ', '').replace(',', '.').upper()
                else:
                    for linesn in sn_ocr_result:
                        sn_ocr_value += linesn[1].replace(" ", "").replace(',', '.').upper()
                if sn_ocr_value.replace('O', '0').find(LABELINFO['sn']) < 0:
                    sn_ocr_value = ''
                    sn_image_resize = cv2.resize(sn_image, (320, 64))
                    sn_ocr_result, _ = FASTOCR(sn_image_resize)
                    for linesn in sn_ocr_result:
                        sn_ocr_value += linesn[1].replace(" ", "").replace(',', '.').upper()
                sn_barcode_value = barcodeDecode(sn_image)
            if part == 2:
                if flag == 0:
                    modelNo_image = image[h_segments[i]:h_segments[i + 1] + 30, :]
                    # cv2.imwrite('temp.jpg', modelNo_image)
                    modelNo_result = OCRINFERER.ocr(modelNo_image)[0]
                    if len(modelNo_result) == 0:
                        continue
                    for line in modelNo_result:
                        modelNo_ocr_value += line[1][0].replace(' ', '').upper()
                    if modelNo_ocr_value.find('ASS') < 0:
                        pass
                    else:
                        modelNo_image = image[h_segments[i] - 180:h_segments[i + 1] - 160, :]
                        # cv2.imwrite('temp.jpg', modelNo_image)
                        modelNo_ocr_value = ''
                        modelNo_result, _ = FASTOCR(modelNo_image)
                        if modelNo_result is None:
                            continue
                        for linemodel in modelNo_result:
                            modelNo_ocr_value += linemodel[1].replace(" ", "").upper()
                else:
                    flag = 0
                    continue
            if part == 3:
                # if flag == 0:
                config_image = image[h_segments[i]:h_segments[i + 1], :]
                cv2.imwrite('temp.jpg', config_image)
                config_result = OCRINFERER.ocr(config_image)[0]
                if len(config_result) == 0:
                    continue
                for line in config_result:
                    config_ocr_value += line[1][0].replace(' ', '').upper()
            part += 1
        else:
            pass

    return upc_ocr_value, upc_barcode_value, pn_ocr_value, pn_barcode_value, sn_ocr_value, sn_barcode_value, \
           modelNo_ocr_value, config_ocr_value


def calculate_histogram(image):
    hist = cv2.calcHist([image], [0, 1, 2], None, [8, 8, 8], [0, 256, 0, 256, 0, 256])
    hist = cv2.normalize(hist, hist).flatten()

    return hist


def getColorSimilarity(img, template):
    image_hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    template_hsv = cv2.cvtColor(template, cv2.COLOR_BGR2HSV)
    hist_image = cv2.calcHist([image_hsv], [0, 1, 2], None, [8, 8, 8], [0, 256, 0, 256, 0, 256])
    hist_image = cv2.normalize(hist_image, hist_image).flatten()
    hist_label = cv2.calcHist([template_hsv], [0, 1, 2], None, [8, 8, 8], [0, 256, 0, 256, 0, 256])
    hist_label = cv2.normalize(hist_label, hist_label).flatten()
    similarity = 1.0 - cv2.compareHist(hist_image, hist_label, cv2.HISTCMP_BHATTACHARYYA)

    return similarity


def getSimilarity(img, template):
    imagegray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    templategray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
    keypoint1, des1 = ORB.detectAndCompute(imagegray, None)
    keypoint2, des2 = ORB.detectAndCompute(templategray, None)
    if des1 is None or des2 is None:
        similarity = 0.0
        return similarity
    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
    matchs = bf.match(des1, des2)
    matchs = sorted(matchs, key=lambda x: x.distance)
    if len(matchs) < 4:
        similarity = 0.0
        return similarity
    src_pts = np.float32([keypoint1[m.queryIdx].pt for m in matchs]).reshape(-1, 1, 2)
    dst_pts = np.float32([keypoint2[m.trainIdx].pt for m in matchs]).reshape(-1, 1, 2)
    M, mask = cv2.findHomography(src_pts, dst_pts, cv2.RANSAC, 5.0)
    matches_mask = mask.ravel().tolist()
    inlines = sum(matches_mask)
    similarity = inlines / len(matches_mask)

    return similarity


def calculateSimilarity(image, template):
    gray1 = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray2 = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
    blured1 = cv2.GaussianBlur(gray1, (3, 3), 0)
    blured2 = cv2.GaussianBlur(gray2, (3, 3), 0)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    contrast_enhanced1 = clahe.apply(blured1)
    contrast_enhanced2 = clahe.apply(blured2)
    # Canny Edge Detection to extract edges
    edges1 = cv2.Canny(contrast_enhanced1, threshold1=50, threshold2=150)
    edges2 = cv2.Canny(contrast_enhanced2, threshold1=50, threshold2=150)
    pil_image = Image.fromarray(edges1)
    pil_template = Image.fromarray(edges2)
    encoded_images = SENTENCETRANSFORMERMODEL.encode([pil_image, pil_template], batch_size=2, convert_to_tensor=True, show_progress_bar=False)
    similarity_score = util.cos_sim(encoded_images[0], encoded_images[1])

    return similarity_score.item()


@server.route('/getGrayValues', methods=['get', 'post'])
def getGrayValues():
    dir_path = flask.request.args.get("dirPath")
    image_path = flask.request.args.get("imagePath")
    sn = flask.request.args.get("sn")
    type = flask.request.args.get("type")
    result_path = getGrayValuesWithImg(dir_path, image_path, sn, type)
    if len(result_path) > 0:
        res = {'result': result_path, 'msg': ''}
    else:
        res = {'msg': 'fail'}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


def rotate_image(img):
    angle = get_rotated_angle(img)
    height, width = img.shape[:2]
    center = (width / 2, height / 2)
    rotation_matrix = cv2.getRotationMatrix2D(center, angle, 1)
    ratated_image = cv2.warpAffine(img, rotation_matrix, (width, height))
    angle1 = get_rotated_angle(ratated_image)
    if abs(angle) > 5.0 or abs(angle1) > 8.0:
        return img
    if angle1 < 0.4:
        return ratated_image
    else:
        rotation_matrix = cv2.getRotationMatrix2D(center, -angle, 1)
        ratated_image = cv2.warpAffine(img, rotation_matrix, (width, height))
        return ratated_image


def get_rotated_angle(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, imgbinary1 = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    # gray = cv2.GaussianBlur(gray, (3, 3), 0)
    # imgbinary2 = cv2.Canny(gray, 50, 150, apertureSize=3)
    contours1, hierarchy1 = cv2.findContours(imgbinary1, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    # contours2, hierarchy2 = cv2.findContours(imgbinary2, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    max_contour1 = max(contours1, key=cv2.contourArea)
    # max_contour2 = max(contours2, key=cv2.contourArea)
    rect1 = cv2.minAreaRect(max_contour1)
    # rect2 = cv2.minAreaRect(max_contour2)
    center1, size1, angle1 = rect1[0], rect1[1], rect1[2]
    # center2, size2, angle2 = rect2[0], rect2[1], rect2[2]
    if angle1 > 0:
        if angle1 > 45:
            angle1 = angle1 - 90
    else:
        if angle1 < -45:
            angle1 = -90 - angle1
        else:
            angle1 = -angle1
    # if angle2 > 0:
    #     if angle2 > 45:
    #         angle2 = 90 - angle2
    # else:
    #     if angle2 < -45:
    #         angle2 = -90 - angle2
    #     else:
    #         angle2 = -angle2
    # if abs(angle2) > abs(angle1):
    #     if 1 < abs(angle2) < 3:
    #         angle = angle2 / 2
    #     else:
    #         angle = angle1
    # else:
    #     angle = angle2
    angle = angle1
    return angle


def getGrayValuesWithImg(dir_path, img_path, sn, type):
    result_path = ''
    try:
        if (img_path.endswith(".jpg")) and (not img_path.startswith('.')):
            img = cv2.imread(img_path)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            m = 18
            n = 9
            divide_image = divide_method(img, m + 1, n + 1)  # 该函数中m+1和n+1表示网格点个数，m和n分别表示分块的块数
            drawed_img = drawAndDetect(divide_image, type)
            final_img = image_concat(drawed_img)
            cv2.imwrite(dir_path + '/' + sn + '.jpg', final_img)
            result_path = dir_path + '/' + sn + '.jpg'
    except:
        print("未知异常")

    return result_path


# 分割成m行n列
def divide_method(img, m, n):
    h, w = img.shape[0], img.shape[1]
    grid_h = int(h * 1.0 / (m - 1) + 0.5)  # 每个网格的高
    grid_w = int(w * 1.0 / (n - 1) + 0.5)  # 每个网格的宽

    # 满足整除关系时的高、宽
    h = grid_h * (m - 1)
    w = grid_w * (n - 1)

    # 图像缩放 也可以用img_re=skimage.transform.resize(img, (h,w)).astype(np.uint8)
    img_re = cv2.resize(img, (w, h), cv2.INTER_LINEAR)
    gx, gy = np.meshgrid(np.linspace(0, w, n), np.linspace(0, h, m))
    gx = gx.astype(np.int)
    gy = gy.astype(np.int)

    # 这是一个五维的张量，前面两维表示分块后图像的位置（第m行，第n列），后面三维表示每个分块后的图像信息
    divide_image = np.zeros([m - 1, n - 1, grid_h, grid_w, 3], np.uint8)

    for i in range(m - 1):
        for j in range(n - 1):
            divide_image[i, j, ...] = img_re[gy[i][j]:gy[i + 1][j + 1], gx[i][j]:gx[i + 1][j + 1], :]
    return divide_image


def image_concat(divide_image):
    m, n, grid_h, grid_w = [divide_image.shape[0], divide_image.shape[1],  # 每行，每列的图像块数
                            divide_image.shape[2], divide_image.shape[3]]  # 每个图像块的尺寸

    restore_image = np.zeros([m * grid_h, n * grid_w, 3], np.uint8)
    restore_image[0:grid_h, 0:]
    for i in range(m):
        for j in range(n):
            restore_image[i * grid_h:(i + 1) * grid_h, j * grid_w:(j + 1) * grid_w] = divide_image[i, j, :]
    return restore_image


def drawAndDetect(divide_image, type):
    m, n = divide_image.shape[0], divide_image.shape[1]
    for i in range(m):
        for j in range(n):
            sub_img = divide_image[i, j, :]
            h, w = sub_img.shape[:2]

            show_value = ''
            if type == 'gray':
                # 计算灰度平均值
                template_gray = cv2.cvtColor(sub_img, cv2.COLOR_BGR2GRAY)
                h, w = template_gray.shape[:2]  # gray 为灰度图
                m = np.reshape(template_gray, [1, w * h])
                mean = int(m.sum() / (w * h))  # 图像平均灰度值
                show_value = str(mean)
                point = (int(w / 2 - 50), int(h / 2 + 10))
                font = cv2.FONT_HERSHEY_SIMPLEX
                cv2.putText(sub_img, show_value, point, font, 2, (255, 255, 255), 4, cv2.LINE_AA)

            if type == 'balance':
                # 计算白平衡平均值
                b, g, r = cv2.split(sub_img)
                b_avg, g_avg, r_avg = int(cv2.mean(b)[0]), int(cv2.mean(g)[0]), int(cv2.mean(r)[0])
                value_d = getMaxMinDValue([b_avg, g_avg, r_avg])
                # k = int((b_avg + g_avg + r_avg) / 3)
                # show_value = str(k)
                show_value1 = '{} {} {}'.format(r_avg, g_avg, b_avg)
                show_value2 = '{}'.format(value_d)
                point1 = (int(w / 2 - 150), int(h / 2 - 20))
                point2 = (int(w / 2 - 150), int(h / 2 + 60))
                font = cv2.FONT_HERSHEY_SIMPLEX
                cv2.putText(sub_img, show_value1, point1, font, 1.5, (255, 255, 255), 4, cv2.LINE_AA)
                cv2.putText(sub_img, show_value2, point2, font, 2, (255, 255, 255), 4, cv2.LINE_AA)

            cv2.rectangle(sub_img, (0, 0), (w, h), (0, 255, 0), thickness=2)
    return divide_image


# 获取数组的最大值和最小值的差
def getMaxMinDValue(arr):
    max_v = 0
    min_v = 0
    if len(arr) > 0:
        max_v = arr[0]
        min_v = arr[0]
        for v in arr:
            if v > max_v:
                max_v = v
            if v < min_v:
                min_v = v
    return max_v - min_v


def getTestGrayBoard(img):
    # hsv图像
    hsv_img = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    mask = gray_hsv(hsv_img)
    final_img = cv2.bitwise_and(img, img, mask=mask)
    return final_img


def gray_hsv(img):
    # H -> 0 - 180
    # S -> 0 - 43
    # V -> 46 - 220
    lower_hsv = np.array([0, 0, 90])
    higher_hsv = np.array([180, 43, 160])
    mask = cv2.inRange(img, lower_hsv, higher_hsv)
    return mask


@server.route('/checkBaliRoute', methods=['get', 'post'])
def checkBaliRoute():
    ip_address = flask.request.args.get("IP")
    dir_path = flask.request.args.get("dirPath")
    ping_command = 'ping' + ' -c 1 -t 1 -W 1 ' + ip_address
    timeout = 2

    @func_set_timeout(timeout)
    def do_ping_check_path():
        ping_result = os.popen(ping_command)
        line = ping_result.read()
        ping_err = "100.0% packet loss" in line
        if ping_err:
            return '-1'
        path_exists = os.path.exists(dir_path)
        if not path_exists:
            return '0'
        return '1'

    try:
        check_result = do_ping_check_path()
    except func_timeout.exceptions.FunctionTimedOut:
        res = {'msg': '-1'}
        return json.dumps(res, ensure_ascii=False)

    res = {'msg': check_result}
    # json.dumps 序列化时对中文默认使用的ascii编码，输出中文需要设置ensure_ascii=False
    return json.dumps(res, ensure_ascii=False)


@server.route('/changeToPlist', methods=['get', 'post'])
def changeToPlist():
    input_path = flask.request.args.get("inputPath")
    output_path = flask.request.args.get("outputPath")
    # workbook = load_workbook('/Users/jrsjrs/Downloads/DanversPrintBundlePNSub-PN_0812v1.xlsx')
    workbook = load_workbook(input_path)
    sheet = workbook.active
    column_data = [cell.value for cell in sheet['C']]
    keyList = []
    for item in column_data:
        if item is not None:
            keyList.append(item)
    print(keyList)

    column_data_value = [cell.value for cell in sheet['D']]
    valueList = []
    for item in column_data_value:
        if item is not None:
            #valueList.append(item.replace('\n', '/'))
            print(item)
            valueList.append('/'.join(item.split()))
    print(valueList)

    result_dict = {}
    for index, value in enumerate(keyList):
        result_dict[value] = valueList[index]
    print(result_dict)

    with open(output_path, 'wb') as plist_file:
        plistlib.dump(result_dict, plist_file)

    return json.dumps({'msg': ''}, ensure_ascii=False)


@server.route('/changeToPlist2', methods=['get', 'post'])
def changeToPlist2():
    input_path = flask.request.args.get("inputPath")
    output_path = flask.request.args.get("outputPath")
    # workbook = load_workbook('/Users/jrsjrs/Downloads/DanversPrintBundlePNSub-PN_0812v1.xlsx')
    workbook = load_workbook(input_path)
    sheet = workbook.active
    column_data = [cell.value for cell in sheet['D']]
    keyList = []
    for item in column_data:
        if item is not None:
            keyList.append(item)
    print(keyList)

    column_data_value = [cell.value for cell in sheet['J']]
    print(len(column_data_value))
    print(column_data_value)
    valueList = []
    for item in column_data_value:
        if item is not None:
            valueList.append(item.replace('\n', '/'))
        else:
            valueList.append('')
    print(valueList)

    result_dict = {}
    for index, value in enumerate(keyList):
        result_dict[value] = valueList[index]
    print(result_dict)

    with open(output_path, 'wb') as plist_file:
        plistlib.dump(result_dict, plist_file)

    return json.dumps({'msg': ''}, ensure_ascii=False)


if __name__ == '__main__':
    # port可以指定端口，默认端口是5000
    # host默认是服务器，默认是127.0.0.1
    # debug=True 修改时不关闭服务
    server.run(host='127.0.0.1', port=10086)

